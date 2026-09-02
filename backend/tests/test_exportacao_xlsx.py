"""
Achado real: os valores monetários saíam nas planilhas exportadas (cotação
e catálogo) sem nenhuma formatação — número cru tipo "253.37" em vez de
"R$ 253,37". Sem HTTP: chama a rota direto (mesmo padrão de
test_editais_filtros.py) e drena o StreamingResponse manualmente pra
inspecionar as células com openpyxl. Rode com:  cd backend && pytest
"""
import asyncio
import io

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.main import cotacao_edital, exportar_produtos
from app.models import Base, Usuario, Edital, ItemEdital, Match, Produto


def _sessao():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    return Session()


def _usuario(db):
    u = Usuario(nome="Teste", email="t@t.com", senha_hash="x")
    db.add(u)
    db.commit()
    return u


def _drenar(response) -> bytes:
    async def _ler():
        partes = []
        async for pedaco in response.body_iterator:
            partes.append(pedaco if isinstance(pedaco, bytes) else pedaco.encode())
        return b"".join(partes)
    return asyncio.run(_ler())


def test_cotacao_xlsx_formata_colunas_de_valor_como_moeda():
    db = _sessao()
    u = _usuario(db)
    ed = Edital(fonte="PNCP", id_externo="ed1", orgao="Orgao Teste",
               objeto="Aquisicao", uf="SP")
    db.add(ed)
    db.commit()
    prod = Produto(usuario_id=u.id, descricao="Papel A4", preco_custo=32.5)
    db.add(prod)
    db.commit()
    item = ItemEdital(edital_id=ed.id, numero=1, descricao="Papel A4 75g",
                      quantidade=10, valor_unitario=50.0)
    db.add(item)
    match = Match(usuario_id=u.id, edital_id=ed.id, score=0.9, nivel="forte",
                  detalhe={"itens": [{"item": 1, "produto_id": prod.id, "confianca": "alta"}]})
    db.add(match)
    db.commit()

    response = cotacao_edital(ed.id, itens=None, fretes=None, user=u, db=db)
    conteudo = _drenar(response)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    # linha 5 é a 1ª linha de item (cabeçalho ocupa até a linha 4)
    for col in ("D", "E", "F", "G"):
        assert ws[f"{col}5"].number_format == "R$ #,##0.00", f"coluna {col} sem formatação de moeda"


def _edital_com_item_e_match(db, u, link_fornecedor="https://fornecedor.exemplo.com/produto/1"):
    ed = Edital(fonte="PNCP", id_externo="ed1", orgao="Orgao Teste",
               objeto="Aquisicao", uf="SP", link="https://pncp.gov.br/app/editais/1/2026/1")
    db.add(ed)
    db.commit()
    prod = Produto(usuario_id=u.id, descricao="Papel A4", preco_custo=32.5,
                   fabricante="Fab", marca="Marca", modelo="Mod",
                   fornecedor_site=link_fornecedor)
    db.add(prod)
    db.commit()
    item = ItemEdital(edital_id=ed.id, numero=1, descricao="Papel A4 75g",
                      quantidade=10, valor_unitario=50.0)
    db.add(item)
    match = Match(usuario_id=u.id, edital_id=ed.id, score=0.9, nivel="forte",
                  detalhe={"itens": [{"item": 1, "produto_id": prod.id, "confianca": "alta"}]})
    db.add(match)
    db.commit()
    return ed, prod


def test_cotacao_xlsx_inclui_coluna_de_link_do_fornecedor_por_item():
    db = _sessao()
    u = _usuario(db)
    ed, prod = _edital_com_item_e_match(db, u)

    response = cotacao_edital(ed.id, itens=None, fretes=None, user=u, db=db)
    wb = openpyxl.load_workbook(io.BytesIO(_drenar(response)))
    ws = wb.active

    assert ws["K4"].value == "LINK"
    # link do PRODUTO (fornecedor_site), não o link do edital no PNCP
    assert ws["K5"].value == prod.fornecedor_site
    assert ws["K5"].value != ed.link
    assert ws["K5"].hyperlink.target == prod.fornecedor_site


def test_cotacao_xlsx_sem_link_de_fornecedor_cadastrado_fica_em_branco():
    db = _sessao()
    u = _usuario(db)
    ed, prod = _edital_com_item_e_match(db, u, link_fornecedor=None)

    response = cotacao_edital(ed.id, itens=None, fretes=None, user=u, db=db)
    wb = openpyxl.load_workbook(io.BytesIO(_drenar(response)))
    ws = wb.active

    assert not ws["K5"].value
    assert ws["K5"].hyperlink is None


def test_cotacao_xlsx_incluir_custo_false_mantem_coluna_mas_deixa_valor_em_branco():
    db = _sessao()
    u = _usuario(db)
    ed, prod = _edital_com_item_e_match(db, u)

    response = cotacao_edital(ed.id, itens=None, fretes=None, incluir_custo=False, user=u, db=db)
    wb = openpyxl.load_workbook(io.BytesIO(_drenar(response)))
    ws = wb.active

    # a coluna continua na planilha (mesma estrutura de sempre) — só o
    # VALOR fica em branco, pra quem recebe a planilha não ver a margem.
    cabecalho = [c.value for c in ws[4]]
    assert cabecalho == ["ITEM", "DESCRIÇÃO", "QTD.", "VALOR UNI.", "VALOR TOTAL",
                          "VALOR MÍNIMO UNI.", "VALOR MÍNIMO TOTAL",
                          "FABRICANTE", "MARCA", "MODELO", "LINK"]
    assert not ws["F5"].value
    assert not ws["G5"].value
    # link não é afetado — continua na mesma coluna de sempre (K)
    assert ws["K4"].value == "LINK"
    assert ws["K5"].value == prod.fornecedor_site


def test_cotacao_xlsx_incluir_custo_true_mantem_comportamento_padrao():
    db = _sessao()
    u = _usuario(db)
    ed, prod = _edital_com_item_e_match(db, u)

    response = cotacao_edital(ed.id, itens=None, fretes=None, user=u, db=db)
    wb = openpyxl.load_workbook(io.BytesIO(_drenar(response)))
    ws = wb.active

    cabecalho = [c.value for c in ws[4]]
    assert cabecalho == ["ITEM", "DESCRIÇÃO", "QTD.", "VALOR UNI.", "VALOR TOTAL",
                          "VALOR MÍNIMO UNI.", "VALOR MÍNIMO TOTAL",
                          "FABRICANTE", "MARCA", "MODELO", "LINK"]
    assert ws["F5"].value == 32.5
    assert ws["G5"].value == "=F5*C5"


def test_catalogo_xlsx_formata_preco_custo_e_venda_como_moeda():
    db = _sessao()
    u = _usuario(db)
    db.add(Produto(usuario_id=u.id, descricao="Papel A4", preco_custo=32.5, preco_venda=45.0))
    db.commit()

    response = exportar_produtos(user=u, db=db)
    conteudo = _drenar(response)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    # linha 2 é a 1ª linha de produto (linha 1 é o cabeçalho); coluna A é o
    # id do produto, então preco_custo/preco_venda ficam em K/L (não J/K).
    assert ws["K2"].number_format == "R$ #,##0.00"
    assert ws["L2"].number_format == "R$ #,##0.00"

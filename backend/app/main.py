"""
API FastAPI do Radar de Licitações + dashboard estático.

Rotas principais:
  GET  /api/produtos            lista catálogo
  POST /api/produtos            cadastra produto
  DEL  /api/produtos/{id}       remove produto
  GET  /api/editais             lista editais com match (filtros: nivel, uf, lido)
  POST /api/editais/{id}/marcar marca lido/interessante
  GET  /api/regras              lista regras de exclusão
  POST /api/regras              cria regra
  DEL  /api/regras/{id}         remove regra
  POST /api/coletar             dispara coleta manual (em background)
  GET  /api/export.csv          exporta matches em CSV
  GET  /api/logs                histórico de coletas
  GET  /api/resumo              KPIs do dashboard
"""
import csv
import io
import os
import base64
import re
import secrets
import threading
import time
import requests
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from urllib.parse import urlparse
from .models import utcnow as _utcnow_main
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Depends, BackgroundTasks, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel
from sqlalchemy import select, func, or_
from sqlalchemy.orm import Session

from .config import settings
from .database import get_session, init_db, SessionLocal
from .models import Produto, Edital, ItemEdital, Match, RegraExclusao, LogColeta, Documento, Proposta, Fornecedor, AnaliseIAExtras
from .service import processar_coleta, podar_editais_orfaos
from .catalogo import catmat


_SECRET_KEY_PADRAO = "troque-isto-em-producao-please-32+chars-aleatorios"


@asynccontextmanager
async def lifespan(app: FastAPI):
    # em produção (APP_BASE_URL https), nunca sobe com a SECRET_KEY padrão do
    # código-fonte: ela assina os tokens de sessão e cifra CPF/CNPJ e chaves
    # de API — com o valor padrão, qualquer um que leia o repositório forja
    # sessão de qualquer usuário. Localmente (sem APP_BASE_URL) segue liberado.
    if settings.APP_BASE_URL.startswith("https") and settings.SECRET_KEY == _SECRET_KEY_PADRAO:
        raise RuntimeError(
            "SECRET_KEY está com o valor padrão do código-fonte. Defina uma "
            "SECRET_KEY própria (variável de ambiente) antes de subir em produção."
        )
    init_db()
    yield


app = FastAPI(title="Minha Licitação", version="2.0", lifespan=lifespan)

# Rotas liberadas sem login (auth, health, cron, página de login e estáticos)
_ROTAS_PUBLICAS = {"/health", "/api/coletar-cron", "/login", "/cadastro", "/verificar", "/redefinir-senha"}
_PREFIXOS_PUBLICOS = ("/api/auth/", "/static/", "/assets/")

BASE_DIR = os.path.dirname(__file__)
# A pasta static fica em backend/static (um nível acima de backend/app)
STATIC_DIR = os.path.join(os.path.dirname(BASE_DIR), "static")


@app.exception_handler(StarletteHTTPException)
async def _pagina_de_erro(request: Request, exc: StarletteHTTPException):
    """Página 404 personalizada pra navegação (link quebrado, URL digitada
    errada). Chamadas de API continuam recebendo JSON — quem consome /api/*
    é o próprio JS do app, que espera {"detail": ...}, não HTML."""
    if exc.status_code == 404 and not request.url.path.startswith("/api/"):
        return FileResponse(os.path.join(STATIC_DIR, "404.html"), status_code=404)
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code,
                        headers=getattr(exc, "headers", None))

BR_TZ = ZoneInfo("America/Sao_Paulo")


def _brt(dt: datetime | None) -> str | None:
    """Converte um datetime UTC (naive) para o horário de Brasília em ISO."""
    if not dt:
        return None
    return dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(BR_TZ).isoformat()


@app.api_route("/health", methods=["GET", "HEAD"])
def health():
    # HEAD explícito: o FastAPI não registra HEAD sozinho pra rotas @app.get,
    # e serviços de keep-alive gratuitos (ex.: UptimeRobot free) só mandam HEAD.
    return {"ok": True}


# =========================== AUTENTICAÇÃO ============================ #
import json as _json_auth
import secrets as _secrets_auth
from email_validator import validate_email, EmailNotValidError
from fastapi import Response as _Resp
from .models import Usuario
from . import auth as _auth
from . import ratelimit as _rl
from .notifications import email as _email_mod


class CadastroIn(BaseModel):
    nome: str
    email: str
    senha: str
    documento: str | None = None       # CPF ou CNPJ
    # complementares (todos opcionais — dá pra completar depois em Meu perfil)
    endereco: dict | None = None        # {cep, logradouro, numero, bairro, cidade, uf, complemento}
    dados_empresa: dict | None = None   # {telefone, representante_legal, inscricao_estadual,
                                         #  inscricao_municipal, banco_nome, banco_agencia, banco_conta}
    logo_base64: str | None = None      # data URI ("data:image/png;base64,...") pra timbrar a proposta


def _email_html_verificacao(nome: str, link: str) -> str:
    """E-mail de confirmação em HTML simples e sóbrio (melhora a entrega)."""
    return f"""\
<!DOCTYPE html>
<html lang="pt-br"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;background:#f4f6f9;font-family:Arial,Helvetica,sans-serif;color:#1a2129">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f9;padding:24px 0">
    <tr><td align="center">
      <table role="presentation" width="480" cellpadding="0" cellspacing="0"
             style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden">
        <tr><td style="background:#14121A;padding:20px 28px;color:#fff;font-size:18px;font-weight:bold">
          Minha Licitação
        </td></tr>
        <tr><td style="padding:28px">
          <p style="margin:0 0 14px;font-size:15px">Olá, {nome}!</p>
          <p style="margin:0 0 20px;font-size:14px;line-height:1.5;color:#3b4654">
            Falta só um passo para ativar a sua conta. Clique no botão abaixo para
            confirmar o seu e-mail.
          </p>
          <p style="margin:0 0 24px;text-align:center">
            <a href="{link}" style="background:#6D28D9;color:#fff;text-decoration:none;
               padding:12px 26px;border-radius:8px;font-size:14px;font-weight:bold;display:inline-block">
              Confirmar meu e-mail
            </a>
          </p>
          <p style="margin:0 0 8px;font-size:12px;color:#5b6770">
            Se o botão não funcionar, copie e cole este endereço no navegador:
          </p>
          <p style="margin:0 0 20px;font-size:12px;color:#6D28D9;word-break:break-all">{link}</p>
          <p style="margin:0;font-size:12px;color:#94a3b8">
            Se você não criou esta conta, é só ignorar esta mensagem.
          </p>
        </td></tr>
      </table>
      <p style="margin:14px 0 0;font-size:11px;color:#94a3b8">Minha Licitação</p>
    </td></tr>
  </table>
</body></html>"""


class LoginIn(BaseModel):
    email: str
    senha: str


def _set_cookie_sessao(resp: _Resp, usuario_id: int):
    """Cookie de um login novo -- sempre abre um teto absoluto novo de
    TOKEN_EXPIRA_HORAS (exp_abs=None). Renovação por atividade dentro de uma
    sessão já aberta é outro caminho (ver auth.get_current_user), que
    carrega o exp_abs original em vez de recriar um."""
    _auth.cookie_sessao(resp, usuario_id)


# campos válidos do JSON cifrado de dados complementares — filtra chave
# desconhecida antes de guardar (o cliente não pode injetar chave arbitrária
# num JSON que vai direto pro banco).
_CAMPOS_DADOS_EMPRESA = {
    "telefone", "representante_legal", "representante_rg", "inscricao_estadual",
    "inscricao_municipal", "banco_nome", "banco_agencia", "banco_conta",
}
# tamanho máximo do data URI da logo (base64 infla ~33% sobre o binário —
# isso cobre uma imagem de até ~800KB, de sobra pra um logo).
_LOGO_MAX_CHARS = 1_200_000


def _limpar_dados_empresa(bruto: dict | None) -> dict:
    if not bruto:
        return {}
    return {k: str(v).strip() for k, v in bruto.items()
           if k in _CAMPOS_DADOS_EMPRESA and str(v or "").strip()}


def _validar_logo_base64(valor: str | None) -> str | None:
    if not valor:
        return None
    valor = valor.strip()
    if not valor.startswith("data:image/"):
        raise HTTPException(400, "Logo inválida — envie uma imagem (PNG, JPG ou SVG).")
    if len(valor) > _LOGO_MAX_CHARS:
        raise HTTPException(400, "Logo muito grande — envie uma imagem menor (até ~800KB).")
    return valor


@app.post("/api/auth/cadastro")
def auth_cadastro(dados: CadastroIn, request: Request, resp: _Resp, bg: BackgroundTasks,
                  db: Session = Depends(get_session)):
    # limite generoso (criar conta não é algo que usuário legítimo faz em
    # rajada) -- barra criação automatizada de contas em massa a partir de
    # um mesmo IP.
    _rl.checar(f"cadastro-ip:{_rl.ip_cliente(request)}", limite=5, janela_seg=3600)
    # valida e-mail
    try:
        email = validate_email(dados.email, check_deliverability=False).normalized.lower()
    except EmailNotValidError:
        raise HTTPException(400, "E-mail inválido.")
    # força da senha
    erro = _auth.validar_forca_senha(dados.senha)
    if erro:
        raise HTTPException(400, erro)
    if not (dados.nome or "").strip():
        raise HTTPException(400, "Informe seu nome.")
    # e-mail único
    existe = db.execute(select(Usuario).where(Usuario.email == email)).scalars().first()
    if existe:
        raise HTTPException(409, "Já existe uma conta com este e-mail.")
    logo = _validar_logo_base64(dados.logo_base64)

    primeiro = db.scalar(select(func.count(Usuario.id))) == 0
    smtp_ok = _email_mod.smtp_configurado()

    import json as _j
    dados_empresa = _limpar_dados_empresa(dados.dados_empresa)
    u = Usuario(
        nome=dados.nome.strip(), email=email,
        senha_hash=_auth.hash_senha(dados.senha),
        doc_cifrado=_auth.cifrar((dados.documento or "").strip() or None),
        endereco_cifrado=_auth.cifrar(_j.dumps(dados.endereco, ensure_ascii=False)) if dados.endereco else None,
        dados_empresa_cifrado=_auth.cifrar(_j.dumps(dados_empresa, ensure_ascii=False)) if dados_empresa else None,
        logo_base64=logo,
        email_verificado=not smtp_ok,   # sem SMTP, libera direto; com SMTP, exige verificar
        token_verificacao=_secrets_auth.token_urlsafe(32) if smtp_ok else None,
    )
    db.add(u)
    db.flush()

    # o primeiro usuário "adota" os dados que já existiam (sem dono)
    if primeiro:
        for tabela in (Produto, Match, Documento, RegraExclusao, Proposta):
            db.query(tabela).filter(tabela.usuario_id.is_(None)).update(
                {tabela.usuario_id: u.id}, synchronize_session=False)

    db.commit()

    if smtp_ok:
        base = settings.APP_BASE_URL.rstrip("/")
        link = f"{base}/verificar?token={u.token_verificacao}"
        corpo = (f"Olá, {u.nome}!\n\nConfirme seu e-mail para ativar a sua conta no "
                 f"Minha Licitação:\n{link}\n\n"
                 "Se você não criou esta conta, ignore esta mensagem.\n\n"
                 "— Minha Licitação")
        html = _email_html_verificacao(u.nome, link)
        # envia em segundo plano: o cadastro responde na hora, sem esperar o e-mail
        bg.add_task(_email_mod.enviar_para, email,
                    "Confirme seu cadastro — Minha Licitação", corpo, html)
        return {"ok": True, "verificar_email": True,
                "mensagem": "Enviamos um link de confirmação para o seu e-mail. "
                            "Confira também a caixa de spam."}

    _set_cookie_sessao(resp, u.id)
    return {"ok": True, "verificar_email": False}


@app.post("/api/auth/login")
def auth_login(dados: LoginIn, request: Request, resp: _Resp, db: Session = Depends(get_session)):
    email = (dados.email or "").strip().lower()
    # duas chaves pegam ataques diferentes: só por IP pega força bruta
    # distribuída entre vários e-mails a partir de uma máquina; só por
    # e-mail pega credential stuffing rotacionando IP contra uma conta só.
    _rl.checar(f"login-ip:{_rl.ip_cliente(request)}", limite=20, janela_seg=300)
    _rl.checar(f"login-email:{email}", limite=6, janela_seg=300)
    u = db.execute(select(Usuario).where(Usuario.email == email)).scalars().first()
    if not u or not _auth.conferir_senha(dados.senha, u.senha_hash):
        raise HTTPException(401, "E-mail ou senha incorretos.")
    if not u.ativo:
        raise HTTPException(403, "Conta desativada.")
    if not u.email_verificado:
        raise HTTPException(403, "Confirme seu e-mail antes de entrar. Verifique sua caixa de entrada.")
    # login deu certo: zera o contador por e-mail -- sem isso, um usuário
    # que só errou a senha uma ou duas vezes antes de acertar ia ficando
    # cada vez mais perto do limite pra sempre, por nenhum motivo real. O
    # contador por IP fica de fora de propósito: ele existe pra pegar
    # varredura contra várias contas, não deve resetar por causa de UM
    # login certo no meio.
    _rl.limpar(f"login-email:{email}")
    _set_cookie_sessao(resp, u.id)
    return {"ok": True}


@app.post("/api/auth/logout")
def auth_logout(resp: _Resp):
    resp.delete_cookie(_auth.COOKIE_NOME, path="/")
    return {"ok": True}


@app.get("/api/auth/me")
def auth_me(user: Usuario = Depends(_auth.get_current_user)):
    return {"id": user.id, "nome": user.nome, "email": user.email,
            "documento": _auth.decifrar(user.doc_cifrado),
            "tem_gemini": bool(user.gemini_key_cifrada),
            "telegram_chat_id": user.telegram_chat_id or "",
            "notif_email": user.notif_email, "notif_telegram": user.notif_telegram}


@app.get("/api/auth/verificar")
def auth_verificar(token: str, db: Session = Depends(get_session)):
    u = db.execute(select(Usuario).where(Usuario.token_verificacao == token)).scalars().first()
    if not u:
        raise HTTPException(400, "Link de verificação inválido ou já usado.")
    u.email_verificado = True
    u.token_verificacao = None
    db.commit()
    return {"ok": True}


class EsqueciSenhaIn(BaseModel):
    email: str


class RedefinirSenhaIn(BaseModel):
    token: str
    senha: str


def _email_html_reset_senha(nome: str, link: str) -> str:
    """E-mail de redefinição de senha em HTML simples e sóbrio."""
    return f"""\
<!DOCTYPE html>
<html lang="pt-br"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;background:#f4f6f9;font-family:Arial,Helvetica,sans-serif;color:#1a2129">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6f9;padding:24px 0">
    <tr><td align="center">
      <table role="presentation" width="480" cellpadding="0" cellspacing="0"
             style="background:#fff;border:1px solid #e2e8f0;border-radius:12px;overflow:hidden">
        <tr><td style="background:#14121A;padding:20px 28px;color:#fff;font-size:18px;font-weight:bold">
          Minha Licitação
        </td></tr>
        <tr><td style="padding:28px">
          <p style="margin:0 0 14px;font-size:15px">Olá, {nome}!</p>
          <p style="margin:0 0 20px;font-size:14px;line-height:1.5;color:#3b4654">
            Recebemos um pedido para redefinir a sua senha. Clique no botão abaixo
            para escolher uma nova senha. Este link expira em 1 hora.
          </p>
          <p style="margin:0 0 24px;text-align:center">
            <a href="{link}" style="background:#6D28D9;color:#fff;text-decoration:none;
               padding:12px 26px;border-radius:8px;font-size:14px;font-weight:bold;display:inline-block">
              Redefinir minha senha
            </a>
          </p>
          <p style="margin:0 0 8px;font-size:12px;color:#5b6770">
            Se o botão não funcionar, copie e cole este endereço no navegador:
          </p>
          <p style="margin:0 0 20px;font-size:12px;color:#6D28D9;word-break:break-all">{link}</p>
          <p style="margin:0;font-size:12px;color:#94a3b8">
            Se você não pediu essa alteração, é só ignorar esta mensagem — sua senha
            continua a mesma.
          </p>
        </td></tr>
      </table>
      <p style="margin:14px 0 0;font-size:11px;color:#94a3b8">Minha Licitação</p>
    </td></tr>
  </table>
</body></html>"""


@app.post("/api/auth/esqueci-senha")
def auth_esqueci_senha(dados: EsqueciSenhaIn, request: Request, bg: BackgroundTasks,
                       db: Session = Depends(get_session)):
    """Sempre responde com sucesso genérico (não revela se o e-mail existe)."""
    mensagem = ("Se este e-mail estiver cadastrado, enviamos um link para redefinir "
                "a senha. Confira também a caixa de spam.")
    email = (dados.email or "").strip().lower()
    # por e-mail: barra mandar o mesmo usuário ser inundado de e-mail de
    # redefinição (achado real conhecido em outros produtos: "e-mail bomb"
    # incomoda a vítima mesmo sem nenhuma conta ser de fato comprometida).
    # por IP: rede mais larga contra varredura de e-mails em massa.
    _rl.checar(f"esqueci-ip:{_rl.ip_cliente(request)}", limite=10, janela_seg=3600)
    if email:
        _rl.checar(f"esqueci-email:{email}", limite=3, janela_seg=3600)
    if not email or not _email_mod.smtp_configurado():
        return {"ok": True, "mensagem": mensagem}

    u = db.execute(select(Usuario).where(Usuario.email == email)).scalars().first()
    if u and u.ativo:
        u.token_reset_senha = _secrets_auth.token_urlsafe(32)
        u.token_reset_expira = _utcnow_main() + timedelta(hours=1)
        db.commit()

        base = settings.APP_BASE_URL.rstrip("/")
        link = f"{base}/redefinir-senha?token={u.token_reset_senha}"
        corpo = (f"Olá, {u.nome}!\n\nRecebemos um pedido para redefinir a sua senha "
                 f"no Minha Licitação. Este link expira em 1 hora:\n{link}\n\n"
                 "Se você não pediu essa alteração, ignore esta mensagem.\n\n"
                 "— Minha Licitação")
        html = _email_html_reset_senha(u.nome, link)
        bg.add_task(_email_mod.enviar_para, email,
                    "Redefinir senha — Minha Licitação", corpo, html)

    return {"ok": True, "mensagem": mensagem}


@app.post("/api/auth/redefinir-senha")
def auth_redefinir_senha(dados: RedefinirSenhaIn, request: Request, db: Session = Depends(get_session)):
    # o token em si já é praticamente impossível de adivinhar (32 bytes
    # aleatórios) -- isto é defesa em profundidade, não a proteção principal.
    _rl.checar(f"redefinir-ip:{_rl.ip_cliente(request)}", limite=10, janela_seg=3600)
    u = db.execute(
        select(Usuario).where(Usuario.token_reset_senha == dados.token)
    ).scalars().first()
    if not u or not u.token_reset_expira or u.token_reset_expira < _utcnow_main():
        raise HTTPException(400, "Link de redefinição inválido ou expirado.")

    erro = _auth.validar_forca_senha(dados.senha)
    if erro:
        raise HTTPException(400, erro)

    u.senha_hash = _auth.hash_senha(dados.senha)
    u.token_reset_senha = None
    u.token_reset_expira = None
    db.commit()
    return {"ok": True}


# =========================== PERFIL ============================ #
class PerfilIn(BaseModel):
    nome: str | None = None
    documento: str | None = None
    gemini_key: str | None = None       # "" limpa; None mantém
    telegram_chat_id: str | None = None
    notif_email: bool | None = None
    notif_telegram: bool | None = None
    avisar_abertura: bool | None = None
    dias_antecedencia: int | None = None
    endereco: dict | None = None        # {cep, logradouro, numero, bairro, cidade, uf, complemento}
    dados_empresa: dict | None = None   # {telefone, representante_legal, inscricao_estadual,
                                         #  inscricao_municipal, banco_nome, banco_agencia, banco_conta}
    logo_base64: str | None = None      # data URI; "" remove a logo, None mantém


@app.get("/api/perfil")
def obter_perfil(user: Usuario = Depends(_auth.get_current_user)):
    import json as _j
    end = _auth.decifrar(user.endereco_cifrado)
    try:
        endereco = _j.loads(end) if end else {}
    except ValueError:
        endereco = {}
    emp = _auth.decifrar(user.dados_empresa_cifrado)
    try:
        dados_empresa = _j.loads(emp) if emp else {}
    except ValueError:
        dados_empresa = {}
    return {
        "nome": user.nome, "email": user.email,
        "documento": _auth.decifrar(user.doc_cifrado) or "",
        "tem_gemini": bool(user.gemini_key_cifrada),
        "telegram_chat_id": user.telegram_chat_id or "",
        "notif_email": user.notif_email, "notif_telegram": user.notif_telegram,
        "avisar_abertura": user.avisar_abertura,
        "dias_antecedencia": user.dias_antecedencia,
        "endereco": endereco,
        "dados_empresa": dados_empresa,
        "logo_base64": user.logo_base64 or "",
    }


@app.post("/api/perfil")
def salvar_perfil(dados: PerfilIn, user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    import json as _j
    if dados.nome is not None and dados.nome.strip():
        user.nome = dados.nome.strip()
    if dados.documento is not None:
        user.doc_cifrado = _auth.cifrar(dados.documento.strip() or None)
    # chave Gemini: None = manter; "" = remover; texto = cifrar e guardar
    if dados.gemini_key is not None:
        user.gemini_key_cifrada = _auth.cifrar(dados.gemini_key.strip() or None)
    if dados.telegram_chat_id is not None:
        user.telegram_chat_id = dados.telegram_chat_id.strip() or None
    if dados.notif_email is not None:
        user.notif_email = dados.notif_email
    if dados.notif_telegram is not None:
        user.notif_telegram = dados.notif_telegram
    if dados.avisar_abertura is not None:
        user.avisar_abertura = dados.avisar_abertura
    if dados.dias_antecedencia is not None:
        user.dias_antecedencia = max(0, min(30, dados.dias_antecedencia))  # 0 a 30 dias
    if dados.endereco is not None:
        user.endereco_cifrado = _auth.cifrar(_j.dumps(dados.endereco, ensure_ascii=False))
    if dados.dados_empresa is not None:
        limpo = _limpar_dados_empresa(dados.dados_empresa)
        user.dados_empresa_cifrado = _auth.cifrar(_j.dumps(limpo, ensure_ascii=False)) if limpo else None
    if dados.logo_base64 is not None:
        user.logo_base64 = _validar_logo_base64(dados.logo_base64) if dados.logo_base64 else None
    db.commit()
    return {"ok": True}


@app.get("/api/cep/{cep}")
def consultar_cep(cep: str, user: Usuario = Depends(_auth.get_current_user)):
    """Autopreenchimento de endereço pelo CEP (ViaCEP, gratuito)."""
    limpo = "".join(c for c in cep if c.isdigit())
    if len(limpo) != 8:
        raise HTTPException(400, "CEP deve ter 8 dígitos.")
    try:
        r = requests.get(f"https://viacep.com.br/ws/{limpo}/json/", timeout=10)
        dados = r.json()
    except Exception:
        raise HTTPException(502, "Não foi possível consultar o CEP agora.")
    if dados.get("erro"):
        raise HTTPException(404, "CEP não encontrado.")
    return {
        "cep": dados.get("cep", ""), "logradouro": dados.get("logradouro", ""),
        "bairro": dados.get("bairro", ""), "cidade": dados.get("localidade", ""),
        "uf": dados.get("uf", ""), "complemento": dados.get("complemento", ""),
    }


# ===================== VÍNCULO DO TELEGRAM (multiusuário) ===================== #
def _campos_telegram(slot: int) -> tuple[str, str]:
    """slot 1 = contato principal, slot 2 = contato adicional (ex.: sócio,
    outro responsável) -- mesmo par código/chat_id, duplicado em Usuario
    pra receber os mesmos avisos de forma independente."""
    if slot == 2:
        return "telegram_codigo_2", "telegram_chat_id_2"
    return "telegram_codigo", "telegram_chat_id"


@app.get("/api/telegram/vinculo")
def telegram_vinculo(slot: int = Query(1), user: Usuario = Depends(_auth.get_current_user),
                     db: Session = Depends(get_session)):
    """Devolve o link para o usuário conectar um Telegram (slot 1 = contato
    principal, slot 2 = um contato adicional) ao bot do Radar. Gera um
    código único na primeira vez."""
    if slot not in (1, 2):
        raise HTTPException(400, "slot inválido")
    campo_codigo, campo_chat = _campos_telegram(slot)
    if not getattr(user, campo_codigo):
        setattr(user, campo_codigo, _secrets_auth.token_urlsafe(8))
        db.commit()
    bot = settings.TELEGRAM_BOT_USERNAME
    disponivel = bool(settings.TELEGRAM_BOT_TOKEN and bot)
    codigo = getattr(user, campo_codigo)
    link = f"https://t.me/{bot}?start={codigo}" if disponivel else ""
    return {
        "disponivel": disponivel,
        "bot": bot,
        "codigo": codigo,
        "link": link,
        "conectado": bool(getattr(user, campo_chat)),
    }


@app.post("/api/telegram/desvincular")
def telegram_desvincular(slot: int = Query(1), user: Usuario = Depends(_auth.get_current_user),
                         db: Session = Depends(get_session)):
    if slot not in (1, 2):
        raise HTTPException(400, "slot inválido")
    _, campo_chat = _campos_telegram(slot)
    setattr(user, campo_chat, None)
    db.commit()
    return {"ok": True}


@app.post("/api/telegram/webhook/{secret}")
async def telegram_webhook(secret: str, req: Request, db: Session = Depends(get_session)):
    """Recebe as mensagens/interações do Telegram. Quando alguém manda
    /start CÓDIGO, vincula o chat_id daquele usuário. Quando toca num botão
    do menu de avisos, manda os editais/documentos da categoria escolhida
    (ver telegram_menu.py). Protegido por um segredo na URL."""
    if not settings.TELEGRAM_WEBHOOK_SECRET or secret != settings.TELEGRAM_WEBHOOK_SECRET:
        raise HTTPException(404, "not found")
    try:
        update = await req.json()
    except Exception:
        return {"ok": True}

    callback = (update or {}).get("callback_query")
    if callback:
        from .notifications import telegram as _tg
        from . import telegram_menu
        callback_id = callback.get("id")
        if callback_id:
            _tg.responder_callback(callback_id)  # tira o "carregando" do botão
        dado = callback.get("data") or ""
        chat_cb = ((callback.get("message") or {}).get("chat") or {})
        chat_id_cb = str(chat_cb.get("id") or "")
        if dado.startswith("radar:") and chat_id_cb:
            categoria = dado.split(":", 1)[1]
            u = db.execute(
                select(Usuario).where(or_(Usuario.telegram_chat_id == chat_id_cb,
                                          Usuario.telegram_chat_id_2 == chat_id_cb))
            ).scalars().first()
            if u and categoria in telegram_menu.CATEGORIAS:
                # manda pro chat que TOCOU no botão (pode ser o 2º contato) --
                # não pro chat_id "principal" fixo, senão a resposta vai pro
                # contato errado quando quem clicou foi o 2º.
                telegram_menu.mostrar_categoria(db, u, categoria, chat_id_cb)
                telegram_menu.enviar_resumo(db, u)  # menu de novo com o que restou
        return {"ok": True}

    msg = (update or {}).get("message") or {}
    texto = (msg.get("text") or "").strip()
    chat = msg.get("chat") or {}
    chat_id = str(chat.get("id") or "")
    if texto.startswith("/start") and chat_id:
        partes = texto.split(maxsplit=1)
        codigo = partes[1].strip() if len(partes) > 1 else ""
        if codigo:
            u = db.execute(select(Usuario).where(Usuario.telegram_codigo == codigo)).scalars().first()
            campo_chat = "telegram_chat_id"
            if not u:
                u = db.execute(select(Usuario).where(Usuario.telegram_codigo_2 == codigo)).scalars().first()
                campo_chat = "telegram_chat_id_2"
            if u:
                setattr(u, campo_chat, chat_id)
                u.notif_telegram = True
                db.commit()
                from .notifications import telegram as _tg
                _tg.enviar_para_chat(
                    chat_id, "✅ Telegram conectado!",
                    f"Pronto, {u.nome}! Você vai receber aqui os avisos de novas "
                    "oportunidades do Minha Licitação.")
                return {"ok": True}
        from .notifications import telegram as _tg
        _tg.enviar_para_chat(
            chat_id, "Minha Licitação",
            "Para conectar, abra o link de vínculo na tela 'Meu perfil' do sistema.")
    return {"ok": True}


@app.post("/api/telegram/registrar-webhook")
def telegram_registrar_webhook(user: Usuario = Depends(_auth.get_current_user)):
    """Registra o webhook no Telegram (rodar uma vez após configurar o bot)."""
    if not (settings.TELEGRAM_BOT_TOKEN and settings.TELEGRAM_WEBHOOK_SECRET and settings.APP_BASE_URL):
        raise HTTPException(400, "Configure TELEGRAM_BOT_TOKEN, TELEGRAM_WEBHOOK_SECRET e APP_BASE_URL.")
    url = f"{settings.APP_BASE_URL.rstrip('/')}/api/telegram/webhook/{settings.TELEGRAM_WEBHOOK_SECRET}"
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/setWebhook",
            json={"url": url, "allowed_updates": ["message", "callback_query"]}, timeout=15)
        return {"ok": r.status_code == 200, "resposta": r.json()}
    except Exception as e:
        raise HTTPException(502, f"Falha ao registrar webhook: {e}")


# --------------------------- Schemas ---------------------------------- #
class ProdutoIn(BaseModel):
    descricao: str
    ncm: str | None = None
    cest: str | None = None
    ean: str | None = None
    catmat: str | None = None
    catser: str | None = None
    palavras_chave: str | None = None
    fabricante: str | None = None
    marca: str | None = None
    modelo: str | None = None
    preco_custo: float | None = None
    preco_venda: float | None = None
    unidade_venda: str | None = None
    itens_por_unidade: float | None = None
    fornecedor_nome: str | None = None
    fornecedor_contato: str | None = None
    fornecedor_site: str | None = None
    fornecedor_id: int | None = None


class RegraIn(BaseModel):
    tipo: str = "termo"
    valor: str


class MarcarIn(BaseModel):
    lido: bool | None = None
    interessante: bool | None = None


class FornecedorIn(BaseModel):
    nome: str
    telefone: str | None = None
    whatsapp: str | None = None
    email: str | None = None
    site: str | None = None
    observacao: str | None = None


def _fornecedor_dict(f: Fornecedor) -> dict:
    return {"id": f.id, "nome": f.nome, "telefone": f.telefone,
            "whatsapp": f.whatsapp, "email": f.email, "site": f.site,
            "observacao": f.observacao}


@app.get("/api/fornecedores")
def listar_fornecedores(user: Usuario = Depends(_auth.get_current_user),
                        db: Session = Depends(get_session)):
    fs = db.execute(select(Fornecedor).where(Fornecedor.usuario_id == user.id,
                    Fornecedor.ativo == True).order_by(Fornecedor.nome.asc())  # noqa: E712
                    ).scalars().all()
    return [_fornecedor_dict(f) for f in fs]


@app.post("/api/fornecedores")
def criar_fornecedor(dados: FornecedorIn, user: Usuario = Depends(_auth.get_current_user),
                     db: Session = Depends(get_session)):
    if not (dados.nome or "").strip():
        raise HTTPException(400, "Informe o nome do fornecedor.")
    f = Fornecedor(**dados.model_dump(), usuario_id=user.id)
    db.add(f)
    db.commit()
    db.refresh(f)
    return _fornecedor_dict(f)


def _fornecedor_do_usuario(db, fid, user) -> Fornecedor:
    f = db.get(Fornecedor, fid)
    if not f or f.usuario_id != user.id:
        raise HTTPException(404, "Fornecedor não encontrado")
    return f


@app.put("/api/fornecedores/{fid}")
def atualizar_fornecedor(fid: int, dados: FornecedorIn,
                         user: Usuario = Depends(_auth.get_current_user),
                         db: Session = Depends(get_session)):
    f = _fornecedor_do_usuario(db, fid, user)
    for campo, valor in dados.model_dump().items():
        setattr(f, campo, valor)
    db.commit()
    return _fornecedor_dict(f)


@app.delete("/api/fornecedores/{fid}")
def remover_fornecedor(fid: int, user: Usuario = Depends(_auth.get_current_user),
                       db: Session = Depends(get_session)):
    f = _fornecedor_do_usuario(db, fid, user)
    f.ativo = False   # soft delete: produtos que apontam pra ele não quebram
    db.commit()
    return {"ok": True}


# --------------------------- Produtos --------------------------------- #
def _produto_dict(p: Produto) -> dict:
    return {
        "id": p.id, "descricao": p.descricao, "ncm": p.ncm, "cest": p.cest,
        "ean": p.ean, "catmat": p.catmat, "catser": p.catser,
        "palavras_chave": p.palavras_chave, "ativo": p.ativo,
        "fabricante": p.fabricante, "marca": p.marca, "modelo": p.modelo,
        "preco_custo": p.preco_custo, "preco_venda": p.preco_venda,
        "unidade_venda": p.unidade_venda, "itens_por_unidade": p.itens_por_unidade,
        "fornecedor_nome": p.fornecedor_nome, "fornecedor_contato": p.fornecedor_contato,
        "fornecedor_site": p.fornecedor_site, "fornecedor_id": p.fornecedor_id,
    }


@app.get("/api/produtos")
def listar_produtos(
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(100, ge=1, le=500),
    user: Usuario = Depends(_auth.get_current_user),
    db: Session = Depends(get_session),
):
    cond = Produto.usuario_id == user.id
    total = db.scalar(select(func.count(Produto.id)).where(cond)) or 0
    produtos = db.execute(
        select(Produto).where(cond).order_by(Produto.id.desc())
        .limit(por_pagina).offset((pagina - 1) * por_pagina)
    ).scalars().all()
    return {
        "total": total, "pagina": pagina, "por_pagina": por_pagina,
        "paginas": (total + por_pagina - 1) // por_pagina,
        "resultados": [_produto_dict(p) for p in produtos],
    }


@app.get("/api/produtos/modelo.xlsx")
def modelo_produtos(user: Usuario = Depends(_auth.get_current_user)):
    """Planilha-modelo para importação de produtos."""
    import openpyxl
    from openpyxl.comments import Comment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"
    cabec = ["descricao", "palavras_chave", "ncm", "ean", "catmat", "catser",
             "fabricante", "marca", "modelo",
             "preco_custo", "preco_venda", "unidade_venda", "itens_por_unidade",
             "fornecedor_telefone", "fornecedor_nome", "fornecedor_contato", "fornecedor_site"]
    ws.append(cabec)
    ws.append(["Papel A4 75g branco", "papel, a4, sulfite, resma", "4802.56.99",
               "7891234567890", "150123", "", "", "", "",
               "18,90", "24,50", "resma", "500",
               "(45) 99999-0000", "", "", ""])
    ws.append(["Caneta esferográfica azul", "caneta, esferográfica, azul", "",
               "", "", "", "", "", "",
               "1,20", "2,00", "unidade", "",
               "", "Distribuidora Exemplo", "(45) 99999-0000", "site.com.br"])
    ws["N1"].comment = Comment(
        "Se o telefone bater com um fornecedor já cadastrado na aba Fornecedores, "
        "o produto é vinculado a ele automaticamente (nome/contato/site vêm de lá "
        "— não precisa preencher as 3 colunas seguintes). Preencha-as só se o "
        "fornecedor ainda não estiver cadastrado.", "Minha Licitação")
    for col in ws.columns:
        larg = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(larg, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_produtos.xlsx"})


@app.get("/api/produtos/exportar.xlsx")
def exportar_produtos(user: Usuario = Depends(_auth.get_current_user),
                      db: Session = Depends(get_session)):
    """Catálogo do usuário em .xlsx, nas MESMAS colunas do modelo de
    importação — reimportar o arquivo exportado (mesmo editado) funciona
    de volta sem ajuste manual."""
    import openpyxl
    produtos = db.execute(
        select(Produto).where(Produto.usuario_id == user.id).order_by(Produto.descricao)
    ).scalars().all()

    # telefone do fornecedor vinculado (se houver) — é o que a importação usa
    # pra religar automaticamente ao mesmo Fornecedor (por isso fica só nessa
    # coluna, não persistido direto no Produto: ver comentário no modelo.xlsx).
    fornecedor_ids = {p.fornecedor_id for p in produtos if p.fornecedor_id}
    telefones = {}
    if fornecedor_ids:
        for f in db.execute(select(Fornecedor).where(Fornecedor.id.in_(fornecedor_ids))).scalars():
            telefones[f.id] = f.telefone or ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"
    cabec = ["descricao", "palavras_chave", "ncm", "ean", "catmat", "catser",
             "fabricante", "marca", "modelo",
             "preco_custo", "preco_venda", "unidade_venda", "itens_por_unidade",
             "fornecedor_telefone", "fornecedor_nome", "fornecedor_contato", "fornecedor_site"]
    ws.append(cabec)
    linha = 1
    for p in produtos:
        ws.append([
            p.descricao, p.palavras_chave, p.ncm, p.ean, p.catmat, p.catser,
            p.fabricante, p.marca, p.modelo,
            p.preco_custo, p.preco_venda, p.unidade_venda, p.itens_por_unidade,
            telefones.get(p.fornecedor_id, ""), p.fornecedor_nome, p.fornecedor_contato, p.fornecedor_site,
        ])
        linha += 1
        # mesmo achado do cotacao.xlsx: preço saía sem formatação de moeda.
        for col in ("J", "K"):
            ws[f"{col}{linha}"].number_format = 'R$ #,##0.00'
    for col in ws.columns:
        larg = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(larg, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = f"catalogo_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"})


def _produto_do_usuario(db, produto_id, user) -> Produto:
    p = db.get(Produto, produto_id)
    if not p or p.usuario_id != user.id:
        raise HTTPException(404, "Produto não encontrado")
    return p


@app.get("/api/produtos/{produto_id}")
def obter_produto(produto_id: int, user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    return _produto_dict(_produto_do_usuario(db, produto_id, user))


@app.post("/api/produtos")
def criar_produto(dados: ProdutoIn, user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    p = Produto(**dados.model_dump(), usuario_id=user.id)
    db.add(p)
    user.versao_catalogo += 1
    db.commit()
    db.refresh(p)
    return {"id": p.id}


@app.put("/api/produtos/{produto_id}")
def atualizar_produto(produto_id: int, dados: ProdutoIn,
                      user: Usuario = Depends(_auth.get_current_user),
                      db: Session = Depends(get_session)):
    p = _produto_do_usuario(db, produto_id, user)
    for campo, valor in dados.model_dump().items():
        setattr(p, campo, valor)
    user.versao_catalogo += 1
    db.commit()
    return {"ok": True, "id": p.id}


# Colunas aceitas na planilha de importação (cabeçalho -> campo do produto)
_COLS_IMPORT = {
    "descricao": "descricao", "descrição": "descricao", "produto": "descricao",
    "palavras_chave": "palavras_chave", "palavras-chave": "palavras_chave",
    "palavras chave": "palavras_chave", "ncm": "ncm", "cest": "cest", "ean": "ean",
    "catmat": "catmat", "catser": "catser",
    "fabricante": "fabricante", "marca": "marca", "modelo": "modelo",
    "preco_custo": "preco_custo", "preço_custo": "preco_custo", "custo": "preco_custo",
    "preco_venda": "preco_venda", "preço_venda": "preco_venda", "venda": "preco_venda",
    "unidade_venda": "unidade_venda", "unidade de venda": "unidade_venda",
    "itens_por_unidade": "itens_por_unidade", "itens por unidade": "itens_por_unidade",
    "fornecedor_nome": "fornecedor_nome", "fornecedor": "fornecedor_nome",
    "fornecedor_contato": "fornecedor_contato", "fornecedor_site": "fornecedor_site",
    # telefone/whatsapp do fornecedor: se bater com um fornecedor já cadastrado,
    # vincula o produto a ele (fornecedor_id) em vez de precisar repetir nome/
    # contato/site em toda linha da planilha.
    "fornecedor_telefone": "_fornecedor_telefone", "telefone_fornecedor": "_fornecedor_telefone",
    "fornecedor_whatsapp": "_fornecedor_telefone", "whatsapp_fornecedor": "_fornecedor_telefone",
}
_CAMPOS_NUM = {"preco_custo", "preco_venda", "itens_por_unidade"}


def _num_br(v):
    """Converte '18,90' / '1.234,56' / 18.9 em float; vazio -> None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s:                       # formato BR: ponto = milhar, vírgula = decimal
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


@app.post("/api/produtos/importar")
async def importar_produtos(arquivo: UploadFile = File(...),
                            user: Usuario = Depends(_auth.get_current_user),
                            db: Session = Depends(get_session)):
    """Importa produtos de uma planilha .xlsx. Atualiza quando a descrição já
    existe; caso contrário, cria. Retorna um resumo do que foi feito."""
    import openpyxl
    conteudo = await arquivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "Arquivo inválido. Envie uma planilha .xlsx.")
    ws = wb.active
    linhas = ws.iter_rows(values_only=True)
    try:
        cabec = next(linhas)
    except StopIteration:
        return {"status": "vazio", "criados": 0, "atualizados": 0, "ignorados": 0, "erros": []}

    # mapeia índice de coluna -> campo do produto
    mapa = {}
    for i, nome in enumerate(cabec):
        chave = str(nome or "").strip().lower()
        if chave in _COLS_IMPORT:
            mapa[i] = _COLS_IMPORT[chave]
    if "descricao" not in mapa.values():
        raise HTTPException(400, "A planilha precisa de uma coluna 'descricao'.")

    # telefone/whatsapp (só dígitos) -> fornecedor já cadastrado deste usuário,
    # para vincular o produto a ele sem precisar repetir nome/contato/site.
    mapa_fone_forn: dict[str, Fornecedor] = {}
    for f in db.execute(select(Fornecedor).where(Fornecedor.usuario_id == user.id)).scalars():
        for fone in (f.telefone, f.whatsapp):
            d = re.sub(r"\D", "", fone or "")
            if d:
                mapa_fone_forn[d] = f

    criados = atualizados = ignorados = 0
    erros = []
    for n, linha in enumerate(linhas, start=2):
        if linha is None or all(c is None or str(c).strip() == "" for c in linha):
            continue
        dados = {}
        for i, campo in mapa.items():
            val = linha[i] if i < len(linha) else None
            if campo in _CAMPOS_NUM:
                dados[campo] = _num_br(val)
            else:
                dados[campo] = (str(val).strip() if val not in (None, "") else None)
        desc = dados.get("descricao")
        if not desc:
            ignorados += 1
            continue
        fone_forn = dados.pop("_fornecedor_telefone", None)
        if fone_forn:
            forn = mapa_fone_forn.get(re.sub(r"\D", "", fone_forn))
            if forn:
                dados["fornecedor_id"] = forn.id
                if not dados.get("fornecedor_nome"):
                    dados["fornecedor_nome"] = forn.nome
                if not dados.get("fornecedor_contato"):
                    dados["fornecedor_contato"] = forn.whatsapp or forn.telefone or forn.email
                if not dados.get("fornecedor_site"):
                    dados["fornecedor_site"] = forn.site
            else:
                erros.append(f"linha {n}: nenhum fornecedor cadastrado com o telefone '{fone_forn}'")
        # atualizar se a descrição já existe NESTE usuário (case-insensitive)
        existente = db.execute(
            select(Produto).where(Produto.usuario_id == user.id)
            .where(func.lower(Produto.descricao) == desc.lower())
        ).scalars().first()
        try:
            if existente:
                for campo, valor in dados.items():
                    if valor is not None:           # só sobrescreve o que veio preenchido
                        setattr(existente, campo, valor)
                atualizados += 1
            else:
                db.add(Produto(**dados, usuario_id=user.id))
                criados += 1
        except Exception as e:
            erros.append(f"linha {n}: {e}")
    if criados or atualizados:
        user.versao_catalogo += 1
    db.commit()
    return {"status": "ok", "criados": criados, "atualizados": atualizados,
            "ignorados": ignorados, "erros": erros[:20]}


def _limpar_matches_do_produto(db: Session, usuario_id: int, produto_id: int) -> None:
    """Remove as referências a um produto excluído dos matches já calculados.
    Sem isso, o card da lista de editais continua mostrando pra sempre o nome
    do produto (texto congelado no momento do match), enquanto a tela de
    detalhes — que confere o produto ao vivo — passa a dizer "sem produto
    compatível" pro mesmo item: as duas telas discordam sobre o mesmo edital.
    Não recalcula o score/nível (isso só fica 100% correto no próximo
    Recalcular); só evita a contradição enquanto isso não roda."""
    matches = db.execute(
        select(Match).where(Match.usuario_id == usuario_id)
    ).scalars().all()
    for m in matches:
        itens = (m.detalhe or {}).get("itens") or []
        if not any(it.get("produto_id") == produto_id for it in itens):
            continue
        restantes = [it for it in itens if it.get("produto_id") != produto_id]
        if not restantes:
            # o match só existia por causa deste produto — sem ele, não é
            # mais uma oportunidade (mesma regra de "não guardamos fracos")
            db.delete(m)
        else:
            m.detalhe = {"itens": restantes}
            m.itens_compativeis = len(restantes)


@app.delete("/api/produtos/{produto_id}")
def remover_produto(produto_id: int, user: Usuario = Depends(_auth.get_current_user),
                    db: Session = Depends(get_session)):
    p = _produto_do_usuario(db, produto_id, user)
    db.delete(p)
    _limpar_matches_do_produto(db, user.id, produto_id)
    user.versao_catalogo += 1
    db.commit()
    return {"ok": True}


class ProdutosIdsIn(BaseModel):
    ids: list[int]


@app.post("/api/produtos/excluir-varios")
def remover_produtos_varios(dados: ProdutosIdsIn,
                            user: Usuario = Depends(_auth.get_current_user),
                            db: Session = Depends(get_session)):
    removidos = 0
    for produto_id in dados.ids:
        p = db.get(Produto, produto_id)
        if not p or p.usuario_id != user.id:
            continue
        db.delete(p)
        _limpar_matches_do_produto(db, user.id, produto_id)
        removidos += 1
    if removidos:
        user.versao_catalogo += 1
    db.commit()
    return {"ok": True, "removidos": removidos}


# --------------------------- Editais / Matches ------------------------ #
def _inicio_hoje_utc() -> datetime:
    """Início do dia de hoje no fuso de Brasília, convertido para UTC naïve
    (coletado_em é gravado em UTC). Serve para contar 'coletados hoje'."""
    tz = ZoneInfo("America/Sao_Paulo")
    agora = datetime.now(tz)
    inicio = agora.replace(hour=0, minute=0, second=0, microsecond=0)
    return inicio.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)


@app.get("/api/editais")
def listar_editais(
    nivel: str | None = Query(None),
    uf: list[str] | None = Query(None),
    status: str | None = Query(None),
    vista: str = Query("ativos", pattern="^(ativos|encerrados|todos)$"),
    apenas_nao_lidos: bool = Query(False),
    apenas_interessantes: bool = Query(False),
    hoje: bool = Query(False),
    tipo: str = Query("todos", pattern="^(todos|produtos|servicos)$"),
    valor_min: float | None = Query(None, ge=0),
    valor_max: float | None = Query(None, ge=0),
    data_de: date | None = Query(None),   # filtra por data_abertura (início de recebimento de propostas)
    data_ate: date | None = Query(None),
    busca_item: str | None = Query(None),
    pagina: int = Query(1, ge=1),
    por_pagina: int = Query(50, ge=1, le=200),
    user: Usuario = Depends(_auth.get_current_user),
    db: Session = Depends(get_session),
):
    hoje_data = date.today()
    base = select(Match, Edital).join(Edital, Match.edital_id == Edital.id)
    filtro = [Match.usuario_id == user.id]
    if nivel:
        filtro.append(Match.nivel == nivel)
    if uf:
        filtro.append(Edital.uf.in_([u.upper() for u in uf]))
    if status:
        filtro.append(Match.status == status)
    if apenas_nao_lidos:
        filtro.append(Match.lido == False)  # noqa: E712
    if apenas_interessantes:
        filtro.append(Match.interessante == True)  # noqa: E712
    if hoje:
        filtro.append(Edital.data_abertura == date.today())
    # tipo: editais que contêm ao menos um item do tipo escolhido (material/serviço)
    if tipo != "todos":
        prefixo = "m" if tipo == "produtos" else "s"
        sub = (select(ItemEdital.edital_id)
               .where(ItemEdital.edital_id == Edital.id)
               .where(func.lower(func.substr(func.coalesce(ItemEdital.material_ou_servico, ""), 1, 1)) == prefixo))
        filtro.append(sub.exists())
    if valor_min is not None:
        filtro.append(Edital.valor_estimado >= valor_min)
    if valor_max is not None:
        filtro.append(Edital.valor_estimado <= valor_max)
    if data_de is not None:
        filtro.append(Edital.data_abertura >= data_de)
    if data_ate is not None:
        filtro.append(Edital.data_abertura <= data_ate)
    # busca por item: só editais que tenham pelo menos um item cujo texto
    # contenha o termo — ex.: usuário digita "grampeador" e só vê os editais
    # que pedem isso, em vez de precisar abrir cada um pra conferir.
    if busca_item and busca_item.strip():
        termo = f"%{busca_item.strip().lower()}%"
        sub_busca = (select(ItemEdital.edital_id)
                    .where(ItemEdital.edital_id == Edital.id)
                    .where(func.lower(ItemEdital.descricao).like(termo)))
        filtro.append(sub_busca.exists())

    if vista == "ativos":
        # por pedido do usuário, "prazo" aqui é a abertura do recebimento de
        # propostas (data_abertura/dataAberturaProposta no PNCP), não o
        # encerramento -- sem data ou data >= hoje conta como ativo
        filtro.append((Edital.data_abertura.is_(None)) |
                      (Edital.data_abertura >= hoje_data))
    elif vista == "encerrados":
        # prazo passou E eu participei (proposta enviada / ganho / perdido)
        filtro.append(Edital.data_abertura < hoje_data)
        filtro.append(Match.status.in_(["proposta_enviada", "ganho", "perdido"]))
    for f in filtro:
        base = base.where(f)

    total = db.scalar(
        select(func.count()).select_from(base.subquery())
    ) or 0

    ordem = (Match.score.desc(), Edital.data_abertura.asc()) if vista == "ativos" \
        else (Edital.data_abertura.desc(),)
    q = base.order_by(*ordem)
    q = q.limit(por_pagina).offset((pagina - 1) * por_pagina)

    from .matching.validacao import validar, classificar

    linhas = db.execute(q).all()
    # itens_por_unidade (campo do catálogo usado por _corrigir_por_embalagem
    # em validacao.py — ver o módulo) não vem "congelado" no detalhe salvo
    # pelo engine.py, só o texto; busca em lote (1 query pra página inteira,
    # não 1 por item) pra não vazar N+1 numa lista de até 200 editais.
    ids_produto: set[int] = set()
    for match, _ed in linhas:
        for it in ((match.detalhe or {}).get("itens") or [])[:4]:
            if it.get("produto_id"):
                ids_produto.add(it["produto_id"])
    itens_por_unidade_map: dict[int, float] = {}
    if ids_produto:
        itens_por_unidade_map = dict(db.execute(
            select(Produto.id, Produto.itens_por_unidade).where(Produto.id.in_(ids_produto))
        ).all())

    def _item_conta_como_compativel(it: dict) -> bool:
        # mesma regra de "compatível de verdade" usada em /detalhe, /cotacao.xlsx
        # e Inteligência de Preço: confiança alta OU confirmado manualmente —
        # confiança média é sugestão, não conta até o usuário confirmar (evita
        # o card da lista anunciar como "compatível" um match que a tela do
        # edital já trata como precisando de confirmação). `"confianca" not in
        # it` é dado ANTERIOR a essa faixa existir (ainda não recalculado) —
        # mantém o comportamento de antes até a próxima recalculação chegar
        # nele, em vez de anunciar "0 compatível" pra base toda de uma vez.
        if not it.get("produto_id"):
            return False
        if "confianca" in it:
            return it.get("confianca") == "alta" or bool(it.get("confirmado_manualmente"))
        return True

    out = []
    for match, ed in linhas:
        dias = (ed.data_abertura - date.today()).days if ed.data_abertura else None
        detalhe = match.detalhe
        itens_compativeis = match.itens_compativeis
        if detalhe and detalhe.get("itens"):
            itens_compativeis = sum(1 for it in detalhe["itens"] if _item_conta_como_compativel(it))
            # cópia (não mutar o dict rastreado pelo ORM) — só os 4 primeiros
            # itens porque é só isso que o card da lista mostra (_corpoItensEdital
            # já não usa esse "detalhe" resumido, usa /detalhe com o catálogo
            # ao vivo). produto/descricao_item já vêm "congelados" no
            # detalhe salvo pelo engine.py, então não precisa buscar Produto
            # no banco aqui — só reaplica a validação em cima do texto já ali.
            itens_copia = [dict(it) for it in detalhe["itens"]]
            for it in itens_copia[:4]:
                if not it.get("produto") or it.get("score_item") is None:
                    continue
                resultado = validar(it.get("descricao_item") or "", it["produto"],
                                    itens_por_unidade_map.get(it.get("produto_id")))
                if resultado.verificavel:
                    it["validacao_tecnica"] = {
                        "classificacao": classificar(it["score_item"], resultado),
                        "criticas": [p.descricao for p in resultado.criticas],
                        "avisos": [p.descricao for p in resultado.avisos],
                    }
            detalhe = {**detalhe, "itens": itens_copia}
        out.append({
            "match_id": match.id, "edital_id": ed.id,
            "orgao": ed.orgao, "objeto": ed.objeto, "uf": ed.uf,
            "municipio": ed.municipio, "modalidade": ed.modalidade,
            "valor_estimado": ed.valor_estimado, "fonte": ed.fonte,
            "data_abertura": ed.data_abertura.isoformat() if ed.data_abertura else None,
            "dias_restantes": dias, "link": ed.link,
            "score": match.score, "nivel": match.nivel,
            "itens_compativeis": itens_compativeis,
            "lido": match.lido, "interessante": match.interessante,
            "status": match.status,
            "detalhe": detalhe,
        })

    # Editais que mencionam o termo buscado mas NUNCA viraram Match (o motor
    # de matching não achou nenhum sinal — comum sem saldo/chave da IA, já
    # que só código fiscal sozinho não vira vencedor automático). Sem isso, a
    # busca por item ficava presa ao mesmo universo "só o que o motor já
    # aprovou", inútil bem na hora em que o motor automático não está
    # disponível. Só computado quando a busca está ativa (evita custo extra
    # em toda listagem normal); limitado a 20 pra não virar outra lista
    # gigante sem paginação.
    sem_match: list[dict] = []
    if busca_item and busca_item.strip():
        termo = f"%{busca_item.strip().lower()}%"
        sub_com_match = select(Match.edital_id).where(Match.usuario_id == user.id)
        q_sem_match = (
            select(Edital)
            .where(~Edital.id.in_(sub_com_match))
            .where(select(ItemEdital.edital_id)
                  .where(ItemEdital.edital_id == Edital.id)
                  .where(func.lower(ItemEdital.descricao).like(termo)).exists())
        )
        if vista == "ativos":
            q_sem_match = q_sem_match.where(
                (Edital.data_abertura.is_(None)) | (Edital.data_abertura >= hoje_data))
        # mesmos filtros de edital aplicados na busca principal (uf, valor,
        # tipo, hoje) — achado real: esta consulta só levava em conta o termo
        # buscado e a "vista", ignorando os demais filtros da tela; resultado
        # era o bloco "sem análise automática" misturando editais de qualquer
        # estado/valor/tipo mesmo com filtros ativos (nivel/status/lido não
        # se aplicam aqui, já que por definição estes editais não têm Match).
        if uf:
            q_sem_match = q_sem_match.where(Edital.uf.in_([u.upper() for u in uf]))
        if tipo != "todos":
            prefixo_sm = "m" if tipo == "produtos" else "s"
            q_sem_match = q_sem_match.where(
                select(ItemEdital.edital_id)
                .where(ItemEdital.edital_id == Edital.id)
                .where(func.lower(func.substr(func.coalesce(ItemEdital.material_ou_servico, ""), 1, 1)) == prefixo_sm)
                .exists())
        if valor_min is not None:
            q_sem_match = q_sem_match.where(Edital.valor_estimado >= valor_min)
        if valor_max is not None:
            q_sem_match = q_sem_match.where(Edital.valor_estimado <= valor_max)
        if hoje:
            q_sem_match = q_sem_match.where(Edital.data_abertura == date.today())
        q_sem_match = q_sem_match.order_by(Edital.coletado_em.desc()).limit(20)
        for ed in db.execute(q_sem_match).scalars().all():
            dias = (ed.data_abertura - date.today()).days if ed.data_abertura else None
            itens_batem = [it.descricao for it in ed.itens
                          if busca_item.strip().lower() in (it.descricao or "").lower()][:3]
            sem_match.append({
                "edital_id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto, "uf": ed.uf,
                "municipio": ed.municipio, "modalidade": ed.modalidade,
                "valor_estimado": ed.valor_estimado,
                "data_abertura": ed.data_abertura.isoformat() if ed.data_abertura else None,
                "dias_restantes": dias, "link": ed.link, "itens_batem": itens_batem,
            })

    return {
        "total": total, "pagina": pagina, "por_pagina": por_pagina,
        "paginas": (total + por_pagina - 1) // por_pagina,
        "resultados": out,
        "sem_match": sem_match,
    }


def _match_do_usuario_por_edital(db, edital_id: int, user: Usuario, nivel_criacao: str = "medio") -> Match:
    """Busca o Match do usuário pra este edital, criando um na hora se não
    existir ainda — mesmo padrão já usado em confirmar_item_edital. Editais
    sem sinal nenhum pro motor automático (nivel "fraco") nunca ganham Match
    sozinhos, mas marcar como lido/interessante ou mudar o status é, por si
    só, sinal suficiente de que o usuário está de olho nesse edital."""
    match = db.execute(select(Match).where(Match.edital_id == edital_id)
                       .where(Match.usuario_id == user.id)).scalar_one_or_none()
    if match:
        return match
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    match = Match(edital_id=edital_id, usuario_id=user.id, score=0.0, nivel=nivel_criacao)
    db.add(match)
    return match


@app.post("/api/editais/{edital_id}/marcar")
def marcar(edital_id: int, dados: MarcarIn,
           user: Usuario = Depends(_auth.get_current_user),
           db: Session = Depends(get_session)):
    m = _match_do_usuario_por_edital(db, edital_id, user)
    if dados.lido is not None:
        m.lido = dados.lido
    if dados.interessante is not None:
        m.interessante = dados.interessante
    db.commit()
    return {"ok": True}


STATUS_VALIDOS = {"novo", "vou_participar", "proposta_enviada", "ganho", "perdido", "descartado"}


class StatusIn(BaseModel):
    status: str


@app.post("/api/editais/{edital_id}/status")
def mudar_status(edital_id: int, dados: StatusIn,
                 user: Usuario = Depends(_auth.get_current_user),
                 db: Session = Depends(get_session)):
    if dados.status not in STATUS_VALIDOS:
        raise HTTPException(400, f"Status inválido. Use um de: {', '.join(sorted(STATUS_VALIDOS))}")
    m = _match_do_usuario_por_edital(db, edital_id, user)
    m.status = dados.status
    # quando o status mudou, não só o valor atual -- alimenta o filtro por
    # mês do card "Editais ganhos" do painel Início (ver GET /api/ganhos).
    m.status_atualizado_em = _utcnow_main()
    db.commit()
    return {"ok": True}


@app.post("/api/editais/{edital_id}/interacao")
def registrar_interacao(edital_id: int,
                        user: Usuario = Depends(_auth.get_current_user),
                        db: Session = Depends(get_session)):
    """Chamado pelo front (silencioso, sem toast) toda vez que o usuário
    navega entre as abas de um edital aberto -- ver abaEdital() no JS.
    Alimenta o card "Analisados recentemente" do painel Início
    (GET /api/editais/recentes)."""
    m = _match_do_usuario_por_edital(db, edital_id, user)
    m.interagido_em = _utcnow_main()
    db.commit()
    return {"ok": True}


def _produto_json(p: Produto) -> dict:
    return {
        "id": p.id, "descricao": p.descricao,
        "preco_custo": p.preco_custo, "preco_venda": p.preco_venda,
        "unidade_venda": p.unidade_venda,
        "itens_por_unidade": p.itens_por_unidade,
        "fornecedor_nome": p.fornecedor_nome,
        "fornecedor_contato": p.fornecedor_contato,
        "fornecedor_site": p.fornecedor_site,
    }


def _qtd_embalagem_pncp(unidade_medida: str | None) -> int | None:
    """Extrai a quantidade de uma embalagem do texto `unidadeMedida` que o
    PNCP manda por item (ex.: "Embalagem 500 FL" -> 500). None quando não
    há número (ex.: "Unidade", "Caixa" sem quantidade, ou campo ausente —
    comum em itens coletados antes desse campo existir)."""
    if not unidade_medida:
        return None
    m = re.search(r"\d+", unidade_medida)
    return int(m.group()) if m else None


# Achado real (edital "PACOTE DE 500 FOLHAS DE PAPEL SULFITE"): o PNCP às
# vezes manda só a ABREVIAÇÃO da unidade, sem número junto (unidadeMedida
# = "PCTE", sem o "500" — diferente do caso "Embalagem 500 FL", que já tem
# o número embutido). _qtd_embalagem_pncp não reconhece esses casos, e o
# cálculo caía de volta pra divisão errada. Lista de unidades do domínio do
# PNCP que representam uma EMBALAGEM/agrupamento (não uma peça avulsa) —
# quando bate uma dessas SEM número, assume mesma base do produto (que já
# é vendido em embalagem, itens_por_unidade > 1), mas sem a certeza que o
# número dá — por isso entra com alerta_unidade=True, pedindo conferência.
_UNIDADES_EMBALAGEM_PNCP = {
    "pacote", "pcte", "pct", "caixa", "cx", "resma", "fardo", "frd",
    "duzia", "dz", "cento", "ct", "rolo", "rl", "kit", "conjunto", "conj",
    "bloco", "bl", "galao", "jogo", "par", "embalagem", "emb",
    "saco", "sacola", "frasco", "pote", "garrafa", "balde", "bombona",
}


def _e_unidade_embalagem_pncp(unidade_medida: str | None) -> bool:
    if not unidade_medida:
        return False
    from .matching.engine import normalizar
    return normalizar(unidade_medida).strip() in _UNIDADES_EMBALAGEM_PNCP


# Achado real (edital "Caneta esferográfica... caixa com 50 unidades", PNCP
# unidadeMedida="CX" sem número): o campo unidadeMedida às vezes vem SÓ com a
# abreviação, mas o tamanho real da embalagem está escrito na DESCRIÇÃO do
# item mesmo — _qtd_embalagem_pncp nunca olhava pra lá, então esses itens
# caíam sempre no ramo "embalagem_nao_confirmada" (alerta_unidade=True) por
# mais que o produto do catálogo já tivesse itens_por_unidade=50 certinho: a
# confirmação era simplesmente impossível de alcançar, não importava o que o
# usuário cadastrasse. Mesma lista de palavras de embalagem, procurando
# "<palavra> [com/de/c/] <número>" na descrição.
_RE_QTD_EMBALAGEM_DESCRICAO = re.compile(
    r"(?:" + "|".join(sorted(_UNIDADES_EMBALAGEM_PNCP, key=len, reverse=True)) +
    r")s?\s*(?:com|de|c)?\s*(\d+)")


def _qtd_embalagem_descricao(descricao: str | None) -> int | None:
    if not descricao:
        return None
    from .matching.engine import normalizar
    # normalizar já baixa a caixa e tira acento/pontuação (então "c/" vira
    # "c" e "caixa," vira "caixa") — a regex roda em cima do texto já limpo.
    m = _RE_QTD_EMBALAGEM_DESCRICAO.search(normalizar(descricao))
    return int(m.group(1)) if m else None


def _custo_e_margem(valor_unitario: float | None, produto: Produto,
                    unidade_medida_item: str | None = None,
                    descricao_item: str | None = None) -> dict:
    """Custo/margem de UM produto contra o valor unitário que o órgão paga
    por um item — mesmo cálculo usado em /detalhe, reaproveitado onde quer
    que a gente precise comparar preço (ex.: sugestão da IA de catálogo).

    Achado real (edital de papel A4): o órgão às vezes já cota o preço na
    MESMA embalagem que o produto do catálogo usa (ex.: R$24,50 por RESMA
    de 500 folhas, igual ao produto vendido em resmas de 500) — dividir o
    custo do catálogo por itens_por_unidade nesse caso comparava preço por
    resma com preço por folha, gerando uma "margem" de 99%+ fictícia (era na
    verdade prejuízo). O PNCP manda o texto da unidade por item
    (`unidadeMedida`, ex.: "Embalagem 500 FL") — quando o número bate com
    itens_por_unidade do produto, o valor já está na mesma base, não divide
    de novo. Quando a unidade é só uma abreviação de embalagem sem número
    (ex.: "PCTE"), tenta achar o tamanho na DESCRIÇÃO do item (ex.: "caixa
    com 50 unidades") antes de desistir — só quando nem isso dá pra achar é
    que assume a mesma coisa com alerta_unidade=True (sem como confirmar o
    TAMANHO, só que não é peça avulsa)."""
    if valor_unitario is None or produto.preco_custo is None:
        return {"margem": None, "margem_pct": None, "custo_comparavel": None, "alerta_unidade": False}
    por_unid = produto.itens_por_unidade if (produto.itens_por_unidade or 0) > 0 else 1
    qtd_embalagem_item = _qtd_embalagem_pncp(unidade_medida_item)
    if qtd_embalagem_item is None:
        qtd_embalagem_item = _qtd_embalagem_descricao(descricao_item)
    embalagem_incompativel = False
    embalagem_nao_confirmada = False
    if por_unid > 1 and qtd_embalagem_item is not None and qtd_embalagem_item == por_unid:
        # órgão já cota por embalagem igual à do produto — mesma base, sem conversão
        custo_comparavel = round(produto.preco_custo, 4)
    elif por_unid > 1 and qtd_embalagem_item is None and _e_unidade_embalagem_pncp(unidade_medida_item):
        custo_comparavel = round(produto.preco_custo, 4)
        embalagem_nao_confirmada = True
    else:
        if por_unid > 1 and qtd_embalagem_item is not None:
            # embalagens de tamanhos DIFERENTES (ex.: item em caixa de 12,
            # produto vendido em pacote de 24) — não dá pra comparar direto
            # com confiança nenhuma das duas formas.
            embalagem_incompativel = True
        custo_comparavel = round(produto.preco_custo / por_unid, 4)
    margem = round(valor_unitario - custo_comparavel, 4)
    margem_pct = round(margem / valor_unitario * 100, 1) if valor_unitario else None
    # se a margem ainda é absurda, provavelmente as unidades não batem
    alerta_unidade = embalagem_incompativel or embalagem_nao_confirmada or (
        margem_pct is not None and (margem_pct < -300 or margem_pct > 300))
    return {"margem": margem, "margem_pct": margem_pct, "custo_comparavel": custo_comparavel,
           "alerta_unidade": alerta_unidade}


def _validacao_tecnica_json(descricao_item: str, produto: Produto, score_semantico: float) -> dict | None:
    """Validação técnica (medidas/material/características) de um produto
    contra a descrição de um item — mesmo mecanismo determinístico usado em
    /detalhe pros itens que o motor de matching já compatibilizou.
    score_semantico: usado só como piso pro classificar() (< 0.35 já mata
    de saída) — quem já vem de um match confiante (código exato, ou uma
    sugestão que outra fonte já validou como relevante) passa 1.0 aqui, pra
    a palavra final ficar 100% com a validação de características (a parte
    que realmente pega os casos de "categoria certa, medida errada")."""
    from .matching.validacao import validar, classificar
    resultado = validar(descricao_item, produto.descricao, produto.itens_por_unidade)
    if not resultado.verificavel:
        return None
    return {
        "classificacao": classificar(score_semantico, resultado),
        "criticas": [p.descricao for p in resultado.criticas],
        "avisos": [p.descricao for p in resultado.avisos],
    }


@app.get("/api/editais/{edital_id}/detalhe")
def edital_detalhe(edital_id: int, user: Usuario = Depends(_auth.get_current_user),
                   db: Session = Depends(get_session)):
    """Detalhes do edital: cada item com o valor pedido pelo órgão, o produto
    (compatível de verdade, ou uma SUGESTÃO a confirmar quando a confiança
    não é alta — ver matching/engine.py), seu preço, a margem e os dados do
    fornecedor."""
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    match = db.execute(select(Match).where(Match.edital_id == edital_id)
                       .where(Match.usuario_id == user.id)).scalar_one_or_none()

    # item (número) -> dado bruto do detalhe do match
    itens_match: dict = {}
    if match and match.detalhe:
        for d in (match.detalhe.get("itens") or []):
            if d.get("item") is not None:
                itens_match[d["item"]] = d

    # busca de uma vez TODOS os produtos referenciados — o selecionado E os
    # candidatos de cada item (pra "sugestoes" trazer dado ao vivo, não só o
    # texto congelado no momento do match)
    prod_ids: set[int] = set()
    for d in itens_match.values():
        if d.get("produto_id"):
            prod_ids.add(d["produto_id"])
        for c in (d.get("candidatos") or []):
            if c.get("produto_id"):
                prod_ids.add(c["produto_id"])
    produtos = {}
    if prod_ids:
        produtos = {p.id: p for p in db.execute(
            select(Produto).where(Produto.id.in_(prod_ids))).scalars()}

    itens = []
    for it in ed.itens:
        d = itens_match.get(it.numero) or {}
        prod = produtos.get(d.get("produto_id"))
        confianca = d.get("confianca")
        confirmado = bool(d.get("confirmado_manualmente"))
        # "compatível" de verdade (entra em cotação/margem/Inteligência de
        # Preço) exige confiança alta OU confirmação manual — confiança
        # média é SUGESTÃO, não vira fato até o usuário confirmar (é
        # exatamente a faixa onde viviam os bugs reais de matching achados
        # em auditoria de produção).
        compativel = prod is not None and (confianca == "alta" or confirmado)

        margem_dados = {"margem": None, "margem_pct": None, "custo_comparavel": None, "alerta_unidade": False}
        validacao_tecnica = None
        if compativel:
            margem_dados = _custo_e_margem(it.valor_unitario, prod, it.unidade_medida, it.descricao)
            # só reporta validação técnica quando havia um score por item
            # disponível — senão fica sem opinião, em vez de inventar um
            # "Atende" sem nenhuma checagem por trás (_validacao_tecnica_json
            # já filtra também por "verificavel" internamente).
            score_item = d.get("score_item")
            if score_item is not None:
                validacao_tecnica = _validacao_tecnica_json(it.descricao, prod, score_item)

        # candidatas com dado ao vivo — sempre presente (mesmo pra item de
        # confiança alta: código NCM/CATMAT exato não é garantia de ser o
        # mesmo produto, o usuário pode querer trocar mesmo esse caso).
        sugestoes = []
        for c in (d.get("candidatos") or []):
            p_c = produtos.get(c.get("produto_id"))
            if p_c:
                sugestoes.append({"score": c.get("score"), "produto": _produto_json(p_c)})

        itens.append({
            "numero": it.numero, "descricao": it.descricao,
            "valor_orgao": it.valor_unitario, "quantidade": it.quantidade,
            "compativel": compativel,
            "confianca": confianca,
            "confirmado_manualmente": confirmado,
            "validacao_tecnica": validacao_tecnica,
            **margem_dados,
            "produto": _produto_json(prod) if (prod and compativel) else None,
            "sugestoes": sugestoes,
        })
    itens.sort(key=lambda x: x["compativel"], reverse=True)

    dias = (ed.data_abertura - date.today()).days if ed.data_abertura else None
    return {
        "edital": {
            "id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto,
            "modalidade": ed.modalidade, "uf": ed.uf, "municipio": ed.municipio,
            "valor_estimado": ed.valor_estimado, "fonte": ed.fonte, "link": ed.link,
            "data_abertura": ed.data_abertura.isoformat() if ed.data_abertura else None,
            "dias_restantes": dias,
            "nivel": match.nivel if match else None,
            "score": match.score if match else None,
            "match_id": match.id if match else None,
            "lido": match.lido if match else None,
            "interessante": match.interessante if match else None,
            "status": match.status if match else None,
            # só indica SE já existe análise em cache — não dispara uma nova
            # (evitar acionar a IA sem o usuário pedir explicitamente).
            "analisado": bool(ed.analise_ia),
            "itens_completados": bool(ed.itens_completados_em),
            "itens_completados_qtd": ed.itens_completados_qtd or 0,
        },
        "itens": itens,
    }


class ConfirmarItemIn(BaseModel):
    produto_id: int | None = None


@app.post("/api/editais/{edital_id}/itens/{numero}/confirmar")
def confirmar_item_edital(edital_id: int, numero: int, body: ConfirmarItemIn,
                          user: Usuario = Depends(_auth.get_current_user),
                          db: Session = Depends(get_session)):
    """Confirma manualmente (ou corrige) o produto do catálogo pra um item do
    edital — tanto pra confirmar uma sugestão de confiança média quanto pra
    trocar um item de confiança alta que o motor errou (código exato não
    garante ser o mesmo produto). `produto_id: null` = "nenhuma destas".
    A confirmação sobrevive a recálculos futuros — ver
    service._mesclar_confirmacoes_manuais."""
    match = db.execute(select(Match).where(Match.edital_id == edital_id)
                       .where(Match.usuario_id == user.id)).scalar_one_or_none()
    if not match:
        # Editais sem nenhum sinal textual (nivel "fraco") nunca ganham Match
        # — de propósito, ver _gerar_matches_usuario. Mas a comparação de
        # catálogo por IA (aba Análise por IA) é independente do motor de
        # texto e pode achar um produto mesmo assim; sem isso, confirmar
        # essa sugestão sempre batia em 404 ("Edital sem match"), mesmo o
        # usuário tendo acabado de confirmar que o produto É o certo — a
        # própria confirmação já é sinal suficiente de relevância.
        ed = db.get(Edital, edital_id)
        if not ed:
            raise HTTPException(404, "Edital não encontrado")
        match = Match(edital_id=edital_id, usuario_id=user.id, score=0.0,
                      nivel="medio", detalhe={"itens": []})
        db.add(match)
    elif not match.detalhe:
        match.detalhe = {"itens": []}

    produto = None
    if body.produto_id is not None:
        produto = _produto_do_usuario(db, body.produto_id, user)

    itens = list(match.detalhe.get("itens") or [])
    idx = next((i for i, d in enumerate(itens) if d.get("item") == numero), None)
    if idx is None:
        # O item pode nunca ter entrado em match.detalhe: o motor de matching
        # só grava um item ali quando o score passa de LIMIAR_ITEM_SUGESTAO
        # (matching/engine.py) — mas a comparação de catálogo por IA (aba
        # Análise por IA) roda uma busca independente e pode sugerir um
        # produto pra um item que o motor descartou por completo. Sem isso,
        # confirmar essa sugestão sempre dava 404 ("Item não encontrado").
        existe = db.execute(select(ItemEdital.id).where(
            ItemEdital.edital_id == edital_id, ItemEdital.numero == numero)).scalar_one_or_none()
        if existe is None:
            raise HTTPException(404, "Item não encontrado neste edital")
        itens.append({"item": numero, "confianca": None, "candidatos": []})
        idx = len(itens) - 1

    item = dict(itens[idx])
    item["produto_id"] = produto.id if produto else None
    item["produto"] = produto.descricao if produto else None
    item["confirmado_manualmente"] = True
    itens[idx] = item
    match.detalhe = {"itens": itens}
    db.commit()
    return {"ok": True}


# --------------------------- Regras de exclusão ----------------------- #
@app.get("/api/regras")
def listar_regras(user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    regras = db.execute(select(RegraExclusao)
                        .where(RegraExclusao.usuario_id == user.id)).scalars().all()
    return [{"id": r.id, "tipo": r.tipo, "valor": r.valor, "ativo": r.ativo} for r in regras]


@app.post("/api/regras")
def criar_regra(dados: RegraIn, user: Usuario = Depends(_auth.get_current_user),
                db: Session = Depends(get_session)):
    r = RegraExclusao(tipo=dados.tipo, valor=dados.valor, usuario_id=user.id)
    db.add(r)
    db.commit()
    return {"id": r.id}


@app.delete("/api/regras/{regra_id}")
def remover_regra(regra_id: int, user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    r = db.get(RegraExclusao, regra_id)
    if r and r.usuario_id == user.id:
        db.delete(r)
        db.commit()
    return {"ok": True}


# --------------------------- Coleta / Logs / Resumo ------------------- #
# Trava para impedir coletas simultâneas (evita condição de corrida que
# duplica matches). Só uma coleta roda por vez; as demais são ignoradas.
_coleta_lock = threading.Lock()
# Uma coleta só (a trava acima é global, não por usuário) — então o
# cancelamento também é global: quem estiver acompanhando pode pedir pra
# parar, não importa quem disparou. Resetada no fim de cada rodada (sucesso,
# erro ou cancelamento) pra não vazar pro próximo disparo.
_coleta_cancelar = False
# Quando a coleta atual começou (None = nenhuma rodando) — usado só pra
# detectar trava PRESA (ver _coleta_travada), não pro indicador de "travado"
# do dashboard (esse já usa iniciado_em do LogColeta, não isto aqui).
_coleta_iniciada_em: datetime | None = None
# Fase atual da coleta em andamento — "buscando" (baixando/gravando editais
# do PNCP) ou "compatibilidade" (calculando compatibilidade pro catálogo de
# cada usuário, ver service.py:processar_coleta). Achado real: o indicador
# do dashboard só dizia "coleta em andamento" do início ao fim — quando a
# busca no PNCP já tinha terminado fazia tempo e só faltava processar
# usuário por usuário, parecia uma trava sem explicação nenhuma.
_coleta_fase: str | None = None
_coleta_fase_feitos = 0
_coleta_fase_total: int | None = None
# Mesmo limiar do indicador "Última coleta não finalizou" (/api/coleta/status)
# — uma coleta de verdade sempre termina bem antes disso (histórico real:
# 50min a ~2h). Achado real: uma coleta manual travou (sem nunca liberar a
# trava — crash ou hang numa chamada de rede sem timeout) e bloqueou TODAS
# as coletas automáticas do cron silenciosamente por horas — o disparo do
# GitHub Actions continuava recebendo 200 (o pedido é só agendado em
# segundo plano), sem nenhum jeito de saber que nada rodou de verdade.
_LIMITE_COLETA_TRAVADA = timedelta(hours=3)


def _coleta_travada() -> bool:
    return (_coleta_lock.locked() and _coleta_iniciada_em is not None
           and (_utcnow_main() - _coleta_iniciada_em) > _LIMITE_COLETA_TRAVADA)


def _limpar_logs_coleta_orfaos(db: Session):
    """Fecha registros de LogColeta que nunca terminaram (processo morto no
    meio — crash, redeploy, ou hang forçado a liberar por _coleta_travada).
    Achado real: 4 rodadas desde jul/2026 ficaram com finalizado_em nulo pra
    sempre — sem isso, o indicador do usuário dono daquele registro fica
    preso mostrando "travado" indefinidamente, mesmo depois da trava em
    memória já ter sumido (um redeploy zera a trava, mas não conserta a
    linha órfã no banco). Roda logo após conseguir a trava, no início de
    uma coleta nova — nesse ponto, qualquer linha ainda aberta É de uma
    rodada anterior morta, nunca desta que está prestes a começar."""
    orfaos = db.execute(
        select(LogColeta).where(LogColeta.finalizado_em.is_(None))
    ).scalars().all()
    if not orfaos:
        return
    agora = _utcnow_main()
    for log_orfao in orfaos:
        log_orfao.erro = "interrompida (processo reiniciado antes de terminar)"
        log_orfao.finalizado_em = agora
    db.commit()


def _atualizar_fase_coleta(fase: str, feitos: int, total: int) -> None:
    global _coleta_fase, _coleta_fase_feitos, _coleta_fase_total
    _coleta_fase, _coleta_fase_feitos, _coleta_fase_total = fase, feitos, total


def _rodar_coleta_bg(usuario_id: int | None = None):
    global _coleta_cancelar, _coleta_iniciada_em, _coleta_fase, _coleta_fase_feitos, _coleta_fase_total
    import logging
    if _coleta_travada():
        # a trava está presa há mais tempo que qualquer coleta legítima
        # levaria — quase certo que o processo anterior morreu (crash, ou
        # hang numa chamada sem timeout) sem nunca chegar no finally que
        # libera. Força a liberação pra não bloquear coletas pra sempre.
        logging.getLogger("coleta").warning(
            "Trava de coleta presa há mais de %s — forçando liberação.", _LIMITE_COLETA_TRAVADA)
        try:
            _coleta_lock.release()
        except RuntimeError:
            pass   # alguém liberou entre a checagem e agora — segue o jogo
        _coleta_iniciada_em = None
    # se já há uma coleta em andamento, não inicia outra (evita duplicatas)
    if not _coleta_lock.acquire(blocking=False):
        logging.getLogger("coleta").info(
            "Coleta já em andamento — ignorando novo disparo.")
        return
    _coleta_iniciada_em = _utcnow_main()
    _coleta_fase, _coleta_fase_feitos, _coleta_fase_total = None, 0, None
    db = SessionLocal()
    try:
        _limpar_logs_coleta_orfaos(db)
        processar_coleta(db, usuario_id=usuario_id, deve_cancelar=lambda: _coleta_cancelar,
                         progresso_fase=_atualizar_fase_coleta)
        # após coletar, verifica prazos encerrando e documentos vencendo
        # (e-mail sai daqui; Telegram é só o resumo com botões, logo abaixo)
        from .lembretes import verificar_todos
        verificar_todos(db)

        import logging as _logging
        from . import telegram_menu
        if usuario_id is not None:
            alvos = [u for u in [db.get(Usuario, usuario_id)] if u]
        else:
            alvos = db.execute(
                select(Usuario).where(Usuario.ativo == True,             # noqa: E712
                                      Usuario.notif_telegram == True,     # noqa: E712
                                      or_(Usuario.telegram_chat_id.is_not(None),
                                          Usuario.telegram_chat_id_2.is_not(None)))
            ).scalars().all()
        for u in alvos:
            try:
                telegram_menu.enviar_resumo(db, u)
            except Exception:
                _logging.getLogger("coleta").exception(
                    "Falha ao enviar resumo do Telegram pro usuário %s", u.id)
    finally:
        db.close()
        _coleta_cancelar = False
        _coleta_iniciada_em = None
        _coleta_fase, _coleta_fase_feitos, _coleta_fase_total = None, 0, None
        # RuntimeError = a trava já tinha sido forçada por outro disparo
        # (ver _coleta_travada) por essa mesma rodada ter passado de 3h —
        # janela rara e aceita: o alternativa (travar pra sempre até um
        # redeploy manual) é bem pior.
        try:
            _coleta_lock.release()
        except RuntimeError:
            pass


@app.post("/api/coletar")
def coletar_agora(bg: BackgroundTasks, user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    # precisa ter produtos cadastrados para a coleta fazer sentido
    tem_produtos = db.scalar(
        select(func.count(Produto.id)).where(Produto.usuario_id == user.id)) or 0
    if not tem_produtos:
        return {"ok": False, "sem_produtos": True,
                "mensagem": "Cadastre ao menos um produto antes de buscar editais."}
    if _coleta_lock.locked():
        return {"ok": False, "em_andamento": True,
                "mensagem": "Já existe uma coleta em andamento (pode ter sido disparada "
                            "por outro usuário ou pelo agendamento automático). Aguarde terminar."}
    # coleta manual gera matches só para quem clicou
    bg.add_task(_rodar_coleta_bg, user.id)
    return {"ok": True, "mensagem": "Coleta iniciada em segundo plano."}


@app.post("/api/coletar/cancelar")
def coletar_cancelar(user: Usuario = Depends(_auth.get_current_user)):
    """Pede pra parar a coleta em andamento (de qualquer usuário — a trava é
    global, só uma roda por vez). Cooperativo: para no próximo ponto de
    checagem (a cada ~100 editais gravados), não instantâneo; o que já foi
    salvo até ali permanece."""
    global _coleta_cancelar
    if not _coleta_lock.locked():
        return {"ok": False, "mensagem": "Nenhuma coleta em andamento."}
    _coleta_cancelar = True
    return {"ok": True, "mensagem": "Cancelamento solicitado — a coleta vai parar em instantes."}


# --------------------------- Recalcular (background) ------------------- #
# Recalcular percorre TODOS os editais já coletados (podem ser dezenas de
# milhares) contra o catálogo. Rodar isso dentro da request HTTP estoura o
# timeout do proxy do Render antes de terminar (o front recebe resposta cortada
# e o "Recalculando" morre na hora). Por isso roda em segundo plano, com o
# mesmo padrão da coleta, e o front acompanha por polling em /api/recalcular/status.
_recalculo_locks: dict[int, threading.Lock] = {}
_recalculo_status: dict[int, dict] = {}
# por usuário (diferente da coleta, que é global) — cada um só pode cancelar
# o PRÓPRIO recálculo, já que a trava de recálculo também é por usuário.
_recalculo_cancelar: dict[int, bool] = {}


def _lock_recalculo(usuario_id: int) -> threading.Lock:
    return _recalculo_locks.setdefault(usuario_id, threading.Lock())


# seletor de modelo do reranker (tela de Recalcular): experimental, por
# enquanto só pra essa conta testar antes de decidir se vale expor geral.
# "gemini" é um PROVEDOR inteiro diferente (usa a chave própria do Gemini do
# usuário, engine.py troca a chamada inteira) — não é um modelo da DeepInfra.
_USUARIOS_SELETOR_RERANKER = {5}
_MODELOS_RERANKER_PERMITIDOS = {
    "Qwen/Qwen3-Reranker-0.6B",
    "Qwen/Qwen3-Reranker-4B",
    "Qwen/Qwen3-Reranker-8B",
    "gemini",
}


def _rodar_recalculo_bg(usuario_id: int, usar_ia: bool | None = None,
                        modelo_reranker: str | None = None):
    import logging
    lock = _lock_recalculo(usuario_id)
    if not lock.acquire(blocking=False):
        return
    db = SessionLocal()
    try:
        from .service import recalcular_matches

        def _prog(feito, total):
            st = _recalculo_status.get(usuario_id, {})
            st.update({"rodando": True, "erro": None, "feito": feito, "total": total})
            _recalculo_status[usuario_id] = st

        resultado = recalcular_matches(
            db, usuario_id=usuario_id, usar_ia=usar_ia, progresso=_prog,
            deve_cancelar=lambda: _recalculo_cancelar.get(usuario_id, False),
            modelo_reranker=modelo_reranker)
        _recalculo_status[usuario_id] = {"rodando": False, "erro": None, **resultado}
    except Exception as e:
        db.rollback()
        # achado real: "log" nunca existiu nesse módulo (main.py sempre usou
        # logging.getLogger(nome) local, nunca um "log" de módulo) — se o
        # recálculo batesse numa exceção inesperada, o PRÓPRIO except quebrava
        # com NameError, mascarando o erro original.
        logging.getLogger("recalculo").exception("Erro ao recalcular matches (usuário %s)", usuario_id)
        _recalculo_status[usuario_id] = {"rodando": False, "erro": str(e)}
    finally:
        db.close()
        _recalculo_cancelar.pop(usuario_id, None)
        lock.release()


@app.post("/api/recalcular")
def recalcular(bg: BackgroundTasks, com_ia: bool = Query(True),
               modelo_reranker: str | None = Query(None),
               user: Usuario = Depends(_auth.get_current_user)):
    """Dispara, em segundo plano, a reavaliação de todos os editais já coletados
    contra o catálogo atual DESTE usuário. Retorna na hora; o resultado é
    consultado em /api/recalcular/status.
    com_ia=false recalcula só por texto (rápido, sem gastar cota).
    modelo_reranker: sobrescreve o modelo padrão da DeepInfra — restrito a
    _USUARIOS_SELETOR_RERANKER e a uma lista fechada de modelos válidos
    (ignorado silenciosamente pra qualquer outra conta ou valor)."""
    lock = _lock_recalculo(user.id)
    if lock.locked():
        return {"ok": False, "em_andamento": True,
                "mensagem": "Já existe um recálculo em andamento."}
    _recalculo_status[user.id] = {"rodando": True, "erro": None, "feito": 0, "total": 0}
    _recalculo_cancelar.pop(user.id, None)   # defensivo: não herdar cancelamento de uma rodada anterior
    # com_ia=True respeita a config; com_ia=False força sem IA
    usar_ia = None if com_ia else False
    modelo = None
    if (modelo_reranker and user.id in _USUARIOS_SELETOR_RERANKER
            and modelo_reranker in _MODELOS_RERANKER_PERMITIDOS):
        modelo = modelo_reranker
    bg.add_task(_rodar_recalculo_bg, user.id, usar_ia, modelo)
    return {"ok": True, "mensagem": "Recálculo iniciado em segundo plano."}


@app.post("/api/recalcular/cancelar")
def recalcular_cancelar(user: Usuario = Depends(_auth.get_current_user)):
    """Pede pra parar o recálculo deste usuário. Cooperativo: para no próximo
    ponto de checagem (a cada ~200 editais), não instantâneo; o que já foi
    salvo até ali permanece."""
    lock = _lock_recalculo(user.id)
    if not lock.locked():
        return {"ok": False, "mensagem": "Nenhum recálculo em andamento."}
    _recalculo_cancelar[user.id] = True
    return {"ok": True, "mensagem": "Cancelamento solicitado — o recálculo vai parar em instantes."}


@app.get("/api/recalcular/status")
def recalcular_status(user: Usuario = Depends(_auth.get_current_user)):
    return _recalculo_status.get(user.id, {"rodando": False, "erro": None})


@app.post("/api/matches/limpar-fracos")
def limpar_fracos(user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    """Remove os matches de baixa compatibilidade ('fracos') do usuário — usado
    uma vez para limpar a base antiga (novas coletas já não salvam fracos)."""
    from sqlalchemy import delete
    res = db.execute(
        delete(Match).where(Match.usuario_id == user.id, Match.nivel == "fraco")
    )
    db.commit()
    return {"ok": True, "removidos": res.rowcount or 0}


def _ref_pncp(ed: Edital):
    """Reconstrói (cnpj, ano, sequencial) a partir do numeroControlePNCP
    (formato: cnpj-tipo-sequencial/ano)."""
    try:
        esq, ano = ed.id_externo.rsplit("/", 1)
        partes = esq.split("-")
        cnpj = (ed.cnpj_orgao or partes[0]).strip()
        seq = int(partes[-1])
        return cnpj, int(ano), seq
    except Exception:
        return None


def _listar_arquivos_pncp(ed: Edital) -> dict:
    """Busca no PNCP os arquivos/anexos publicados para o edital."""
    ref = _ref_pncp(ed)
    if not ref:
        return {"status": "sem_ref", "arquivos": [], "portal": ed.link}
    cnpj, ano, seq = ref
    base = settings.PNCP_ITENS_BASE_URL.rstrip("/")
    url = f"{base}/v1/orgaos/{cnpj}/compras/{ano}/{seq}/arquivos"
    try:
        r = requests.get(url, timeout=30,
                         headers={"Accept": "application/json",
                                  "User-Agent": "RadarLicitacoes/1.0"})
    except requests.RequestException:
        return {"status": "erro_rede", "arquivos": [], "portal": ed.link}
    if r.status_code != 200:
        return {"status": f"http_{r.status_code}", "arquivos": [], "portal": ed.link}
    try:
        dados = r.json()
    except ValueError:
        return {"status": "resposta_invalida", "arquivos": [], "portal": ed.link}

    lista = dados if isinstance(dados, list) else (dados.get("data") or [])
    arquivos = []
    for a in lista:
        if not isinstance(a, dict):
            continue
        seq_doc = a.get("sequencialDocumento")
        arquivos.append({
            "titulo": a.get("titulo") or a.get("nomeArquivo")
                      or a.get("tipoDocumentoNome") or "Documento",
            "tipo": a.get("tipoDocumentoNome") or "",
            "url": a.get("url") or a.get("uri") or a.get("link")
                   or (f"{url}/{seq_doc}" if seq_doc is not None else None),
        })
    arquivos = [x for x in arquivos if x["url"]]
    return {"status": "ok" if arquivos else "vazio",
            "arquivos": arquivos, "portal": ed.link}


def _backfill_unidade_medida(ed: Edital) -> int:
    """Busca no PNCP a unidadeMedida de cada item do edital e preenche nos
    que ainda estão sem (achado real: esse campo não era capturado antes —
    ver _custo_e_margem — causando cálculo de margem errado quando o órgão
    já cota o preço na mesma embalagem do produto do catálogo). Idempotente
    (só toca item com unidade_medida None) e best-effort: falha de rede/PNCP
    não quebra, só não atualiza nada nesse edital. Retorna quantos itens
    foram atualizados — NÃO comita, quem chama decide quando."""
    ref = _ref_pncp(ed)
    if not ref:
        return 0
    cnpj, ano, seq = ref
    base = settings.PNCP_ITENS_BASE_URL.rstrip("/")
    url = f"{base}/v1/orgaos/{cnpj}/compras/{ano}/{seq}/itens"
    try:
        r = requests.get(url, params={"pagina": 1, "tamanhoPagina": 100}, timeout=30,
                         headers={"Accept": "application/json", "User-Agent": "RadarLicitacoes/1.0"})
    except requests.RequestException:
        return 0
    if r.status_code != 200:
        return 0
    try:
        dados = r.json()
    except ValueError:
        return 0
    lista = dados if isinstance(dados, list) else (dados.get("data") or [])
    unidades_por_numero = {it.get("numeroItem"): it.get("unidadeMedida")
                           for it in lista if isinstance(it, dict)}
    atualizados = 0
    for item in ed.itens:
        if item.unidade_medida is not None:
            continue
        nova = unidades_por_numero.get(item.numero)
        if nova:
            item.unidade_medida = nova
            atualizados += 1
    return atualizados


_backfill_unidade_status: dict = {"rodando": False, "feito": 0, "total": 0,
                                  "itens_atualizados": 0, "erro": None}


def _rodar_backfill_unidade_bg():
    import logging
    db = SessionLocal()
    try:
        # mesmo filtro de "edital relevante" usado em _gerar_matches_usuario:
        # ativo (prazo em aberto) OU encerrado mas com algum usuário
        # acompanhando (proposta enviada/ganho/perdido). Sem isso, o backfill
        # perdia tempo com o histórico inteiro (dezenas de milhares de
        # editais encerrados que ninguém mais vê) em vez de só os que ainda
        # importam pro cálculo de margem de alguém.
        hoje = date.today()
        sub_acompanhados = select(Match.edital_id).where(
            Match.status.in_(("proposta_enviada", "ganho", "perdido")))
        eds = db.execute(
            select(Edital)
            .where(Edital.itens.any(ItemEdital.unidade_medida.is_(None)))
            .where(
                (Edital.data_abertura.is_(None))
                | (Edital.data_abertura >= hoje)
                | (Edital.id.in_(sub_acompanhados))
            )
        ).scalars().unique().all()
        _backfill_unidade_status.update(
            {"rodando": True, "feito": 0, "total": len(eds), "itens_atualizados": 0, "erro": None})
        for i, ed in enumerate(eds):
            n = _backfill_unidade_medida(ed)
            _backfill_unidade_status["itens_atualizados"] += n
            _backfill_unidade_status["feito"] = i + 1
            if (i + 1) % 50 == 0:
                db.commit()
            time.sleep(settings.PNCP_DELAY)
        db.commit()
    except Exception as e:
        logging.getLogger("backfill_unidade").exception("Erro no backfill de unidade_medida")
        _backfill_unidade_status["erro"] = str(e)
    finally:
        _backfill_unidade_status["rodando"] = False
        db.close()


@app.post("/api/admin/backfill-unidade-medida")
def backfill_unidade_medida_endpoint(bg: BackgroundTasks, edital_id: int | None = Query(None),
                                     user: Usuario = Depends(_auth.get_current_user),
                                     db: Session = Depends(get_session)):
    """Busca no PNCP a unidadeMedida de itens já coletados antes desse campo
    existir (ver _custo_e_margem). Com edital_id: roda na hora, só nesse
    edital (rápido). Sem edital_id: roda em segundo plano pra TODOS os
    editais com algum item sem essa informação (1 chamada ao PNCP por
    edital, pode demorar bastante numa base grande — acompanhar em
    /api/admin/backfill-unidade-medida/status)."""
    if edital_id is not None:
        ed = db.get(Edital, edital_id)
        if not ed:
            raise HTTPException(404, "Edital não encontrado")
        n = _backfill_unidade_medida(ed)
        db.commit()
        return {"editais_atualizados": 1 if n else 0, "itens_atualizados": n}
    if _backfill_unidade_status["rodando"]:
        return {"ok": False, "mensagem": "Já tem um backfill rodando."}
    bg.add_task(_rodar_backfill_unidade_bg)
    return {"ok": True, "mensagem": "Backfill iniciado em segundo plano."}


@app.get("/api/admin/backfill-unidade-medida/status")
def backfill_unidade_medida_status(user: Usuario = Depends(_auth.get_current_user)):
    return _backfill_unidade_status


_poda_editais_status: dict = {"rodando": False, "editais_removidos": 0, "itens_removidos": 0, "erro": None}


def _rodar_poda_editais_bg():
    import logging
    db = SessionLocal()
    try:
        _poda_editais_status.update(
            {"rodando": True, "editais_removidos": 0, "itens_removidos": 0, "erro": None})
        r = podar_editais_orfaos(db)
        _poda_editais_status.update(r)
    except Exception as e:
        logging.getLogger("poda_editais").exception("Erro ao podar editais órfãos")
        _poda_editais_status["erro"] = str(e)
    finally:
        _poda_editais_status["rodando"] = False
        db.close()


@app.post("/api/admin/podar-editais-orfaos")
def podar_editais_orfaos_endpoint(bg: BackgroundTasks, user: Usuario = Depends(_auth.get_current_user)):
    """Faxina pontual pra reclamar espaço agora (a poda automática -- ver
    service.podar_editais_orfaos, chamada ao fim de toda coleta -- já evita
    a base crescer de novo, mas não limpa o que já se acumulou antes dela
    existir). Roda em segundo plano (base grande pode levar um tempo);
    acompanhar em /api/admin/podar-editais-orfaos/status."""
    if _poda_editais_status["rodando"]:
        return {"ok": False, "mensagem": "Já tem uma poda rodando."}
    bg.add_task(_rodar_poda_editais_bg)
    return {"ok": True, "mensagem": "Poda iniciada em segundo plano."}


@app.get("/api/admin/podar-editais-orfaos/status")
def podar_editais_orfaos_status(user: Usuario = Depends(_auth.get_current_user)):
    return _poda_editais_status


@app.get("/api/editais/{edital_id}/documentos")
def documentos_edital(edital_id: int, user: Usuario = Depends(_auth.get_current_user),
                      db: Session = Depends(get_session)):
    """Lista os arquivos/anexos do edital publicados no PNCP para download."""
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    return _listar_arquivos_pncp(ed)


@app.get("/api/documentos/baixar")
def baixar_documento_pncp(url: str, nome: str = Query("documento"),
                          user: Usuario = Depends(_auth.get_current_user)):
    """Repassa (proxy) um arquivo do PNCP com Content-Disposition: attachment.
    Baixar VÁRIOS documentos de uma vez abrindo uma aba por arquivo esbarra no
    bloqueador de pop-up do navegador (só a 1ª aba de um clique conta como
    gesto do usuário — as demais são bloqueadas). Como este endpoint força
    download em vez de abrir aba, o front pode disparar todos em sequência
    sem esbarrar nisso. Só aceita URLs do domínio do PNCP (evita virar um
    proxy aberto para qualquer endereço)."""
    partes = urlparse(url)
    host = partes.hostname or ""
    if partes.scheme not in ("http", "https") or not (host == "pncp.gov.br" or host.endswith(".pncp.gov.br")):
        raise HTTPException(400, "URL inválida.")
    try:
        r = requests.get(url, stream=True, timeout=60,
                         headers={"User-Agent": "RadarLicitacoes/1.0"})
    except requests.RequestException:
        raise HTTPException(502, "Não foi possível baixar o arquivo do PNCP.")
    if r.status_code != 200:
        raise HTTPException(502, f"PNCP retornou HTTP {r.status_code}.")
    nome_seguro = re.sub(r'[\\/:*?"<>|\r\n]', "_", nome).strip()[:150] or "documento"
    nome_seguro = nome_seguro.encode("latin-1", errors="ignore").decode("latin-1") or "documento"
    # o "nome" vem do título do documento no PNCP (ex.: "Edital"), sem extensão.
    # Descobre a extensão certa pelo nome de arquivo original do PNCP (se ele
    # mandar um Content-Disposition) ou, na falta dele, pelo Content-Type.
    ext = ""
    cd_origem = r.headers.get("Content-Disposition", "")
    m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd_origem)
    if m and "." in m.group(1):
        ext = "." + m.group(1).rsplit(".", 1)[-1].strip('"')
    if not ext:
        import mimetypes
        ext = mimetypes.guess_extension((r.headers.get("Content-Type") or "").split(";")[0].strip()) or ""
    if ext and not nome_seguro.lower().endswith(ext.lower()):
        nome_seguro += ext
    return StreamingResponse(
        r.iter_content(chunk_size=65536),
        media_type=r.headers.get("Content-Type", "application/octet-stream"),
        headers={"Content-Disposition": f'attachment; filename="{nome_seguro}"'})


def _anexar_checklist_documentos(resultado: dict, user: Usuario, db: Session) -> dict:
    """Cruza os documentos de habilitação exigidos (já na análise) com os
    documentos que o usuário tem cadastrados. Roda em toda leitura (nunca faz
    parte do cache da análise) porque validade de documento muda todo dia —
    uma análise de 3 dias atrás não pode continuar dizendo "válido" pra um
    documento que venceu ontem."""
    if resultado.get("status") != "ok":
        return resultado
    docs_usuario = db.execute(
        select(Documento).where(Documento.usuario_id == user.id, Documento.ativo == True)  # noqa: E712
    ).scalars().all()
    from . import checklist_habilitacao
    resultado["checklist_documentos"] = checklist_habilitacao.montar(
        resultado.get("documentos_habilitacao") or {},
        [{"id": d.id, "nome": d.nome, "data_validade": d.data_validade, "ativo": d.ativo}
         for d in docs_usuario],
    )
    return resultado


def _obter_cache_extras(db: Session, user: Usuario, edital_id: int) -> "AnaliseIAExtras | None":
    return db.execute(select(AnaliseIAExtras).where(
        AnaliseIAExtras.usuario_id == user.id, AnaliseIAExtras.edital_id == edital_id
    )).scalars().first()


def _upsert_cache_extras(db: Session, user: Usuario, edital_id: int, cache: "AnaliseIAExtras | None",
                         *, campo_valor: str, valor_json: str | None,
                         campo_versao: str, versao: int) -> None:
    if not cache:
        cache = AnaliseIAExtras(usuario_id=user.id, edital_id=edital_id)
        db.add(cache)
    setattr(cache, campo_valor, valor_json)
    setattr(cache, campo_versao, versao)
    db.commit()


# Mesmo problema do _analise_locks (Edital.analise_ia), só que na cache POR
# USUÁRIO (AnaliseIAExtras, única por usuario_id+edital_id): duas requests
# do MESMO usuário pro MESMO edital ao mesmo tempo (duplo clique, 2 abas)
# liam a cache como inexistente as duas e tentavam INSERIR a mesma linha —
# violava a constraint de unicidade (achado real, via teste de concorrência
# da trava acima). Trava por (usuario_id, edital_id) fecha a janela entre
# "ler se já existe" e "criar se não existe" — usada tanto pela verificação
# de documentos quanto pela comparação de catálogo abaixo (são chamadas em
# sequência, nunca em paralelo, dentro da MESMA request — só serializa
# entre requests diferentes).
_extras_locks: dict[tuple[int, int], threading.Lock] = {}


def _lock_extras(usuario_id: int, edital_id: int) -> threading.Lock:
    return _extras_locks.setdefault((usuario_id, edital_id), threading.Lock())


def _anexar_verificacao_ia_documentos(resultado: dict, ed: Edital, user: Usuario, db: Session,
                                      api_key: str | None, forcar: bool = False) -> dict:
    """Verificação por CONTEÚDO (IA) dos documentos que o usuário já tem
    cadastrados (aba Documentos, cada um com o texto extraído do arquivo
    anexado no cadastro) contra o que este edital exige — complementar ao
    checklist por NOME (_anexar_checklist_documentos, sempre grátis e
    instantâneo, roda em toda leitura).

    É por usuário (não por edital, ao contrário de Edital.analise_ia), então
    fica cacheada à parte em AnaliseIAExtras, versionada por
    Usuario.versao_documentos: só chama a IA de novo quando nunca foi
    calculada pra este edital, quando o usuário pediu explicitamente
    (forcar=True, botão "Realizar nova análise") ou quando a versão mudou —
    e mesmo nesse último caso, devolve o resultado antigo (com a flag
    "verificacao_documentos_desatualizada") em vez de gastar uma chamada de
    IA sem o usuário ter pedido; achado real: gerava uma chamada de IA a
    cada abertura da aba, mesmo sem nada ter mudado no catálogo/documentos."""
    if resultado.get("status") != "ok":
        return resultado
    with _lock_extras(user.id, ed.id):
        return _verificar_ia_documentos_com_cache(resultado, ed, user, db, api_key, forcar)


def _verificar_ia_documentos_com_cache(resultado: dict, ed: Edital, user: Usuario, db: Session,
                                       api_key: str | None, forcar: bool = False) -> dict:
    import json as _json
    cache = _obter_cache_extras(db, user, ed.id)
    tem_cache = cache is not None and cache.versao_documentos_calc is not None
    if tem_cache and not forcar:
        if cache.verificacao_documentos_ia:
            resultado["verificacao_documentos_ia"] = _json.loads(cache.verificacao_documentos_ia)
        if cache.versao_documentos_calc != user.versao_documentos:
            resultado["verificacao_documentos_desatualizada"] = True
        return resultado
    docs_usuario = db.execute(
        select(Documento).where(Documento.usuario_id == user.id, Documento.ativo == True,  # noqa: E712
                                Documento.texto_extraido.is_not(None))
    ).scalars().all()
    saida = None
    if docs_usuario:
        from . import analise_edital as ia
        saida = ia.verificar_documentos_usuario(
            resultado.get("objeto") or "", resultado.get("requisitos_tecnicos"),
            resultado.get("documentos_habilitacao"),
            [{"nome": d.nome, "texto": d.texto_extraido} for d in docs_usuario],
            api_key=api_key,
        )
    _upsert_cache_extras(db, user, ed.id, cache,
        campo_valor="verificacao_documentos_ia",
        valor_json=_json.dumps(saida, ensure_ascii=False) if saida else None,
        campo_versao="versao_documentos_calc", versao=user.versao_documentos)
    if saida:
        resultado["verificacao_documentos_ia"] = saida
    return resultado


def _anexar_comparacao_catalogo_ia(resultado: dict, ed: Edital, user: Usuario, db: Session,
                                   api_key: str | None, forcar: bool = False, deve_cancelar=None) -> dict:
    """Segunda opinião da IA: manda o CATÁLOGO COMPLETO do usuário e os itens
    deste edital pra comparação direta — independente do motor de matching
    por texto (matching/engine.py). Mesma regra de cache versionado do
    _anexar_verificacao_ia_documentos, só que por Usuario.versao_catalogo."""
    if resultado.get("status") != "ok":
        return resultado
    with _lock_extras(user.id, ed.id):
        return _comparar_catalogo_ia_com_cache(resultado, ed, user, db, api_key, forcar, deve_cancelar)


def _comparar_catalogo_ia_com_cache(resultado: dict, ed: Edital, user: Usuario, db: Session,
                                    api_key: str | None, forcar: bool = False, deve_cancelar=None) -> dict:
    import json as _json
    cache = _obter_cache_extras(db, user, ed.id)
    tem_cache = cache is not None and cache.versao_catalogo_calc is not None
    if tem_cache and not forcar:
        if cache.comparacao_catalogo_ia:
            resultado["comparacao_catalogo_ia"] = _json.loads(cache.comparacao_catalogo_ia)
        if cache.versao_catalogo_calc != user.versao_catalogo:
            resultado["comparacao_catalogo_desatualizada"] = True
        return resultado
    itens_edital = db.execute(
        select(ItemEdital).where(ItemEdital.edital_id == ed.id)).scalars().all()
    catalogo = db.execute(
        select(Produto).where(Produto.usuario_id == user.id)).scalars().all()
    if not itens_edital or not catalogo:
        _upsert_cache_extras(db, user, ed.id, cache,
            campo_valor="comparacao_catalogo_ia", valor_json=None,
            campo_versao="versao_catalogo_calc", versao=user.versao_catalogo)
        return resultado
    from . import analise_edital as ia
    saida = ia.comparar_catalogo_usuario(
        resultado.get("objeto") or ed.objeto or "",
        [{"numero": it.numero, "descricao": it.descricao} for it in itens_edital],
        [{"id": p.id, "descricao": p.descricao} for p in catalogo],
        api_key=api_key, deve_cancelar=deve_cancelar,
    )
    if saida.get("status") == "ok":
        # anexa dado ao vivo do produto/item pra tabela do front não precisar
        # de outra chamada — a IA só devolveu numero_item/produto_id. Também
        # roda a MESMA validação técnica determinística (medidas/material/
        # características) usada pros itens que o motor de matching já
        # compatibilizou — achado real: a IA sugeriu "Fita ... 48mm x 50m"
        # pra um item que pedia "largura: 50mm" — mesma categoria de
        # produto, medida diferente, e a IA sozinha não percebeu. score=1.0
        # porque quem decide "atende"/"não atende" aqui é a validação de
        # características, não mais um score textual (a IA já vouched pela
        # relevância geral).
        produtos_map = {p.id: p for p in catalogo}
        itens_map = {it.numero: it for it in itens_edital}
        enriquecidos = []
        for it in saida["itens"]:
            ie = itens_map.get(it["numero"])
            if not ie:
                continue
            candidatos = []
            for c in it["candidatos"]:
                p = produtos_map.get(c["produto_id"])
                if not p:
                    continue
                candidatos.append({
                    "produto_id": p.id, "produto": _produto_json(p),
                    "justificativa": c["justificativa"],
                    "validacao_tecnica": _validacao_tecnica_json(ie.descricao, p, 1.0),
                    **_custo_e_margem(ie.valor_unitario, p, ie.unidade_medida, ie.descricao),
                })
            if not candidatos:
                continue
            enriquecidos.append({
                "numero": it["numero"], "descricao_item": ie.descricao,
                "valor_orgao": ie.valor_unitario, "quantidade": ie.quantidade,
                "candidatos": candidatos,
            })
        saida["itens"] = enriquecidos
    _upsert_cache_extras(db, user, ed.id, cache,
        campo_valor="comparacao_catalogo_ia", valor_json=_json.dumps(saida, ensure_ascii=False),
        campo_versao="versao_catalogo_calc", versao=user.versao_catalogo)
    resultado["comparacao_catalogo_ia"] = saida
    return resultado


# Análise por IA não roda em BackgroundTasks — ao contrário de coleta/
# recálculo, é uma request síncrona comum, sem "processo de fundo" pra ter
# uma trava. Só a flag de cancelamento, por usuário, checada ENTRE as até 3
# chamadas de IA que uma análise faz (edital -> documentos -> catálogo).
# Nunca interrompe uma chamada já em voo — requests.post dentro de
# analise_edital.py:_gerar() não tem nenhum hook de cancelamento.
_analise_cancelar: dict[int, bool] = {}
# Achado real: Edital.analise_ia é cache GLOBAL por edital (o resumo é o
# mesmo pra todo mundo, não depende do usuário) — mas a rota não tinha
# nenhuma trava. Dois usuários abrindo o MESMO edital ainda sem análise ao
# mesmo tempo disparavam a chamada de IA em dobro (a 2ª sobrescrevia o
# resultado da 1ª sem corromper nada, só pagava a chamada à toa). Trava por
# edital_id (mesmo padrão de _recalculo_locks, por usuário) resolve: quem
# chega depois espera aqui e, ao entrar, relê o cache — se o primeiro já
# terminou, aproveita o resultado dele em vez de chamar a IA de novo.
_analise_locks: dict[int, threading.Lock] = {}


def _lock_analise(edital_id: int) -> threading.Lock:
    return _analise_locks.setdefault(edital_id, threading.Lock())


@app.post("/api/editais/{edital_id}/analise/cancelar")
def analise_cancelar(edital_id: int, user: Usuario = Depends(_auth.get_current_user)):
    """Pede pra parar a análise em andamento deste usuário. Cooperativo: só
    tem efeito na PRÓXIMA etapa (documentos/catálogo) ou, se pedido antes de
    a análise-base começar, evita gastar essa chamada — nunca interrompe uma
    chamada à IA já em curso."""
    _analise_cancelar[user.id] = True
    return {"ok": True, "mensagem": "Cancelamento solicitado — a análise vai parar assim que possível."}


def _rodar_extras_ia(resultado: dict, ed: Edital, user: Usuario, db: Session,
                     api_key: str | None, deve_cancelar, forcar: bool = False) -> dict:
    """Roda verificação de documentos + comparação de catálogo, checando
    cancelamento ENTRE as duas etapas (nunca no meio de uma chamada já em
    voo). Usado tanto no cache hit quanto numa análise fresca — mesma
    checagem nos dois lugares. `forcar` (== o ?forcar=true da rota, botão
    "Realizar nova análise") é o único jeito de recalcular as duas checagens
    quando o catálogo/documentos mudaram — sem isso, uma mudança só marca o
    resultado anterior como desatualizado (ver _anexar_verificacao_ia_documentos
    e _anexar_comparacao_catalogo_ia), nunca gasta uma chamada de IA sozinha."""
    if deve_cancelar():
        resultado["cancelado"] = True
        return resultado
    resultado = _anexar_verificacao_ia_documentos(resultado, ed, user, db, api_key, forcar)
    if deve_cancelar():
        resultado["cancelado"] = True
        return resultado
    return _anexar_comparacao_catalogo_ia(resultado, ed, user, db, api_key, forcar, deve_cancelar)


@app.get("/api/editais/{edital_id}/analise")
def analise_edital(edital_id: int, forcar: bool = Query(False),
                   user: Usuario = Depends(_auth.get_current_user),
                   db: Session = Depends(get_session)):
    """Análise do edital por IA (resumo, exigências, prazos, pontos de atenção).
    Resultado fica em cache; use ?forcar=true para refazer."""
    from . import analise_edital as ia
    import json as _json
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    chave = _auth.decifrar(user.gemini_key_cifrada)
    _analise_cancelar.pop(user.id, None)   # defensivo: não herdar cancelamento de uma rodada anterior
    deve_cancelar = lambda: _analise_cancelar.get(user.id, False)  # noqa: E731

    def _cache_valido():
        if not ed.analise_ia:
            return None
        try:
            cache = _json.loads(ed.analise_ia)
        except ValueError:
            return None
        return cache if cache.get("versao") == ia.VERSAO_PROMPT else None

    # análise já feita: mostra do cache (é leitura, não consome IA pro texto
    # do edital em si), desde que tenha sido gerada com a versão atual do
    # prompt. Versão antiga -> refaz. Verificação de documentos/comparação de
    # catálogo são cacheadas à parte, por usuário, em AnaliseIAExtras — só
    # recalculam (gastando uma chamada de IA) na 1ª vez pra este edital ou
    # quando forcar=True; se o catálogo/documentos mudaram nesse meio tempo,
    # o resultado anterior volta marcado como desatualizado em vez de ser
    # refeito sem o usuário pedir (ver _anexar_verificacao_ia_documentos e
    # _anexar_comparacao_catalogo_ia).
    cache = None if forcar else _cache_valido()
    if cache:
        cache["cache"] = True
        cache = _rodar_extras_ia(cache, ed, user, db, chave, deve_cancelar, forcar)
        return _anexar_checklist_documentos(cache, user, db)
    # para RODAR uma análise nova, exige a chave Gemini do próprio usuário
    if not ia.ia_texto_disponivel(chave):
        return {"status": "sem_ia"}
    if deve_cancelar():
        return {"status": "cancelado"}
    # trava por edital: o cache (ed.analise_ia) é global, não por usuário —
    # sem isso, dois usuários abrindo o MESMO edital ainda sem análise ao
    # mesmo tempo disparavam a chamada de IA em dobro. Quem chega enquanto
    # outro já está gerando espera aqui; ao entrar, relê o cache (db.refresh
    # — a outra request usa outra sessão) e aproveita o resultado dela se já
    # tiver terminado, em vez de pagar a chamada de novo.
    with _lock_analise(edital_id):
        db.refresh(ed)
        resultado = None if forcar else _cache_valido()
        if resultado:
            resultado["cache"] = True
        else:
            docs = _listar_arquivos_pncp(ed)
            resultado = ia.analisar(ed.objeto or "", docs.get("arquivos") or [], api_key=chave)
            if resultado.get("status") == "ok":
                ed.analise_ia = _json.dumps(resultado, ensure_ascii=False)
                ed.analise_em = datetime.now(ZoneInfo("America/Sao_Paulo")).replace(tzinfo=None)
                db.commit()
    resultado = _rodar_extras_ia(resultado, ed, user, db, chave, deve_cancelar, forcar)
    return _anexar_checklist_documentos(resultado, user, db)


# Achado real: essa leitura (baixar o documento + calcular embeddings +
# chamar a IA, com retentativa se a DeepInfra estiver lenta) pode passar de
# 100-160s em editais grandes — rodando direto na request HTTP, o proxy do
# Render derruba a conexão com 502 antes de terminar (mesmo motivo do
# recálculo, ver comentário lá em cima de _recalculo_locks). O trabalho
# continuava rodando no servidor e salvando no final, mas o navegador nunca
# via o resultado a tempo — parecia que "não tinha feito nada". Mesmo
# padrão de segundo plano + polling do recálculo, só que por EDITAL (não
# por usuário): ItemEdital.descricao é compartilhada entre todo mundo que
# vê este edital, então a trava também é.
_completar_descricao_locks: dict[int, threading.Lock] = {}
_completar_descricao_status: dict[int, dict] = {}


def _lock_completar_descricao(edital_id: int) -> threading.Lock:
    return _completar_descricao_locks.setdefault(edital_id, threading.Lock())


def _rodar_completar_descricao_bg(edital_id: int):
    import logging
    from . import itens_pdf
    lock = _lock_completar_descricao(edital_id)
    if not lock.acquire(blocking=False):
        return
    db = SessionLocal()
    try:
        ed = db.get(Edital, edital_id)
        if not ed:
            _completar_descricao_status[edital_id] = {"rodando": False, "erro": "Edital não encontrado"}
            return
        if not itens_pdf.ia_disponivel(settings.DEEPINFRA_API_KEY):
            _completar_descricao_status[edital_id] = {
                "rodando": False, "erro": None, "status": "sem_ia", "atualizados": 0}
            return

        itens_edital = db.execute(
            select(ItemEdital).where(ItemEdital.edital_id == edital_id)).scalars().all()
        docs = _listar_arquivos_pncp(ed)
        resultado = itens_pdf.extrair_itens_completos(
            ed.objeto or "", docs.get("arquivos") or [],
            [{"numero": it.numero, "descricao": it.descricao} for it in itens_edital],
        )
        atualizados = 0
        if resultado.get("status") == "ok":
            mapa = {it.numero: it for it in itens_edital}
            for r in resultado["itens"]:
                alvo = mapa.get(r["numero"])
                if alvo and r["descricao_completa"] and r["descricao_completa"] != alvo.descricao:
                    logging.getLogger("itens_pdf").info(
                        "Item %s do edital %s completado via PDF (%d -> %d chars)",
                        r["numero"], edital_id, len(alvo.descricao or ""), len(r["descricao_completa"]))
                    alvo.descricao = r["descricao_completa"]
                    atualizados += 1
            # só marca "completado" (o que impede o disparo automático de
            # tentar de novo ao abrir a aba) quando a tentativa realmente
            # terminou com um resultado -- achado real: antes disso era
            # gravado sempre, mesmo em "sem_texto"/"erro_ia"/etc., e um
            # edital que falhasse por qualquer motivo (documento escaneado,
            # instabilidade passageira da IA) nunca mais tentava de novo
            # sozinho, mesmo que uma tentativa futura pudesse dar certo.
            ed.itens_completados_em = datetime.now(ZoneInfo("America/Sao_Paulo")).replace(tzinfo=None)
            # achado real: itens_completados_em setado não significa que
            # algo foi melhorado, só que a tentativa terminou -- guarda
            # quantos itens essa tentativa atualizou de verdade, pra tela
            # não anunciar "descrições completadas" quando não completou nada.
            ed.itens_completados_qtd = atualizados
        db.commit()
        _completar_descricao_status[edital_id] = {
            "rodando": False, "erro": None, "status": resultado.get("status"),
            "detalhe": resultado.get("detalhe"), "atualizados": atualizados}
    except Exception as e:
        db.rollback()
        logging.getLogger("itens_pdf").exception("Erro ao completar descrição de itens (edital %s)", edital_id)
        _completar_descricao_status[edital_id] = {"rodando": False, "erro": str(e)}
    finally:
        db.close()
        lock.release()


@app.post("/api/editais/{edital_id}/itens/completar-descricao")
def completar_descricao_itens(edital_id: int, bg: BackgroundTasks,
                              user: Usuario = Depends(_auth.get_current_user),
                              db: Session = Depends(get_session)):
    """Dispara em segundo plano a leitura do(s) documento(s) do edital
    publicados no PNCP pra tentar completar a descrição de itens que a API
    estruturada trouxe cortada — usa uma IA paga pelo OPERADOR (DeepInfra,
    ver app/itens_pdf.py), não a chave Gemini pessoal do usuário. Sobrescreve
    ItemEdital.descricao quando encontra o texto completo com confiança;
    precisa de recálculo depois pra refletir no motor de matching. Retorna
    na hora; o resultado é consultado em .../completar-descricao/status."""
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    lock = _lock_completar_descricao(edital_id)
    if lock.locked():
        return {"ok": False, "em_andamento": True,
                "mensagem": "Já tem uma busca de descrição em andamento pra este edital."}
    _completar_descricao_status[edital_id] = {"rodando": True, "erro": None}
    bg.add_task(_rodar_completar_descricao_bg, edital_id)
    return {"ok": True, "em_andamento": True}


@app.get("/api/editais/{edital_id}/itens/completar-descricao/status")
def completar_descricao_status(edital_id: int, user: Usuario = Depends(_auth.get_current_user)):
    return _completar_descricao_status.get(edital_id, {"rodando": False, "erro": None})


# --------------------------- Cotação (planilha) ------------------------ #
_MODALIDADE_ABREV = {
    "pregao eletronico": "PE", "pregao presencial": "PP",
    "concorrencia eletronica": "CE", "concorrencia presencial": "CP",
    "dispensa": "DISPENSA", "inexigibilidade": "INEXIG.",
    "credenciamento": "CREDENC.", "leilao eletronico": "LE",
    "leilao presencial": "LP", "dialogo competitivo": "DC",
    "concurso": "CONCURSO", "manifestacao de interesse": "MIP",
    "pre qualificacao": "PRÉ-QUALIF.",
}


def _numero_processo_pncp(ed: Edital) -> str:
    """Monta "Nº sequencial/ano" a partir do numeroControlePNCP (via
    _ref_pncp, já usado pra buscar documentos) — dado estruturado, sempre
    confiável quando o edital veio do PNCP. `ed.raw` NÃO serve pra isso: o
    coletor zera esse campo antes de salvar (evita inflar o banco), então
    ficaria sempre vazio pra qualquer edital coletado normalmente."""
    ref = _ref_pncp(ed)
    if ref:
        _, ano, seq = ref
        return f"{seq}/{ano}"
    return ""


def _linha_cabecalho_cotacao(ed: Edital, analise: dict | None) -> str:
    from .matching.engine import normalizar
    abrev = _MODALIDADE_ABREV.get(normalizar(ed.modalidade or ""), (ed.modalidade or "").upper())
    numero = _numero_processo_pncp(ed)
    orgao_dados = (analise or {}).get("dados_orgao") or {}
    plataforma = orgao_dados.get("plataforma") or ""
    data_sessao = orgao_dados.get("data_sessao") or (
        ed.data_abertura.strftime("%d/%m/%Y") if ed.data_abertura else "")
    partes = [p for p in ([f"Nº {numero}"] if numero else []) + [plataforma, data_sessao] if p]
    return " - ".join([abrev] + partes) if abrev else " - ".join(partes)


@app.get("/api/editais/{edital_id}/cotacao.xlsx")
def cotacao_edital(edital_id: int, itens: str | None = Query(None),
                   fretes: str | None = Query(None),
                   user: Usuario = Depends(_auth.get_current_user),
                   db: Session = Depends(get_session)):
    """Planilha de cotação (mesmo modelo usado no dia do pregão): só os itens
    compatíveis com o catálogo, com fabricante/marca/modelo e valor mínimo
    (custo cadastrado + frete de entrada/saída, informados por ITEM — cada
    item pode ter vindo de um fornecedor diferente, então não dá pra ratear
    um frete único entre eles). Alguns campos do cabeçalho (plataforma,
    horário da sessão) só saem preenchidos se a análise por IA já tiver
    rodado pra este edital — o resto funciona sem ela.
    `itens`: números separados por vírgula (ex.: "3,4") — o usuário escolhe
    quais itens compatíveis entram na planilha; sem o parâmetro, entram
    todos (mantém o link antigo funcionando).
    `fretes`: JSON {"<numero_item>": {"entrada": valor_total, "saida": valor_total}}
    — valor TOTAL do frete daquele item (fornecedor→você e você→órgão),
    informado na hora da cotação; dividido pela quantidade do item pra virar
    custo por unidade. Não fica salvo em lugar nenhum."""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    import json as _json

    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")

    match = db.execute(select(Match).where(Match.edital_id == edital_id)
                       .where(Match.usuario_id == user.id)).scalar_one_or_none()
    mapa_produto: dict[int, int] = {}
    if match and match.detalhe:
        for d in (match.detalhe.get("itens") or []):
            # só entra na cotação o que é confiável de fato (código exato /
            # score alto) ou que o usuário já confirmou manualmente — item
            # de confiança média ainda não confirmado é só uma sugestão.
            if (d.get("item") is not None and d.get("produto_id")
                    and (d.get("confianca") == "alta" or d.get("confirmado_manualmente"))):
                mapa_produto[d["item"]] = d["produto_id"]

    prod_ids = set(mapa_produto.values())
    produtos = {}
    if prod_ids:
        produtos = {p.id: p for p in db.execute(
            select(Produto).where(Produto.id.in_(prod_ids))).scalars()}

    numeros_selecionados: set[int] | None = None
    if itens:
        try:
            numeros_selecionados = {int(n) for n in itens.split(",") if n.strip()}
        except ValueError:
            raise HTTPException(400, "Parâmetro 'itens' inválido — use números separados por vírgula.")

    itens_edital = db.execute(
        select(ItemEdital).where(ItemEdital.edital_id == edital_id)
        .order_by(ItemEdital.numero.asc())).scalars().all()
    linhas = [(it, produtos[mapa_produto[it.numero]]) for it in itens_edital
             if it.numero in mapa_produto and mapa_produto[it.numero] in produtos
             and (numeros_selecionados is None or it.numero in numeros_selecionados)]
    if not linhas:
        raise HTTPException(400, "Nenhum item selecionado bate com o seu catálogo — não há o que cotar.")

    fretes_por_item: dict[int, dict[str, float]] = {}
    if fretes:
        try:
            bruto = _json.loads(fretes)
            for chave, valores in bruto.items():
                fretes_por_item[int(chave)] = {
                    "entrada": float((valores or {}).get("entrada") or 0),
                    "saida": float((valores or {}).get("saida") or 0),
                }
        except (ValueError, TypeError, AttributeError, KeyError):
            raise HTTPException(400, "Parâmetro 'fretes' inválido.")

    analise = None
    if ed.analise_ia:
        try:
            analise = _json.loads(ed.analise_ia)
        except ValueError:
            analise = None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotação"
    negrito = Font(bold=True)
    quebra = Alignment(wrap_text=True, vertical="top")

    ws.append([f"CLIENTE: {ed.orgao or ''}  CNPJ: {ed.cnpj_orgao or ''}"])
    ws["A1"].font = negrito
    ws.append([_linha_cabecalho_cotacao(ed, analise)])
    ws.append([])
    cabec = ["ITEM", "DESCRIÇÃO", "QTD.", "VALOR UNI.", "VALOR TOTAL",
             "VALOR MÍNIMO UNI.", "VALOR MÍNIMO TOTAL", "FABRICANTE", "MARCA", "MODELO"]
    ws.append(cabec)
    for c in ws[4]:
        c.font = negrito

    linha = 5
    for it, prod in linhas:
        qtd = it.quantidade or 0.0
        frete_item = fretes_por_item.get(it.numero, {})
        frete_unit = ((frete_item.get("entrada", 0.0) + frete_item.get("saida", 0.0)) / qtd) if qtd else 0.0
        custo_com_frete = round((prod.preco_custo or 0.0) + frete_unit, 4)
        ws.append([
            it.numero, it.descricao, it.quantidade,
            it.valor_unitario, f"=D{linha}*C{linha}",
            custo_com_frete, f"=F{linha}*C{linha}",
            prod.fabricante, prod.marca, prod.modelo,
        ])
        ws.cell(row=linha, column=2).alignment = quebra
        # achado real: valores saíam sem formatação nenhuma (número cru tipo
        # "253.37" em vez de "R$ 253,37") — aplica formato de moeda nas 4
        # colunas de valor (as duas com fórmula também: number_format afeta
        # só a exibição do resultado calculado, não o cálculo em si).
        for col in ("D", "E", "F", "G"):
            ws[f"{col}{linha}"].number_format = 'R$ #,##0.00'
        linha += 1

    linha += 1  # linha em branco antes das notas
    proposta = (analise or {}).get("dados_proposta") or {}
    notas = []
    if proposta.get("validade_dias"):
        notas.append(f"PROPOSTA: {proposta['validade_dias']}")
    if proposta.get("prazo_entrega"):
        notas.append(f"ENTREGA: {proposta['prazo_entrega']}")
    if proposta.get("garantia_produto"):
        notas.append(f"GARANTIA: {proposta['garantia_produto']}")
    for nota in notas:
        ws.cell(row=linha, column=2, value=nota)
        linha += 1

    # pontos de atenção: NÃO é texto fixo — vem da análise por IA, uma
    # cláusula específica DESTE edital (ex.: exigência de catálogo/atestado
    # anexado à proposta, declaração de Anexo VI, etc.). Sem análise, essa
    # parte simplesmente não aparece (mesma regra de plataforma/data_sessao).
    pontos_atencao = (analise or {}).get("pontos_atencao") or []
    if notas or pontos_atencao:
        linha += 1
    for ponto in pontos_atencao:
        cel = ws.cell(row=linha, column=2, value=ponto)
        cel.alignment = quebra
        linha += 2

    larguras = {"A": 8, "B": 50, "C": 8, "D": 12, "E": 12,
               "F": 14, "G": 14, "H": 16, "I": 16, "J": 16}
    for col, larg in larguras.items():
        ws.column_dimensions[col].width = larg

    numero = _numero_processo_pncp(ed).replace("/", "-")
    nome_arquivo = f"Cotacao_{numero}.xlsx" if numero else f"Cotacao_edital_{edital_id}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'})


@app.api_route("/api/coletar-cron", methods=["GET", "POST"])
def coletar_cron(bg: BackgroundTasks, request: Request):
    """Dispara a coleta de DENTRO do Render (que alcança o PNCP), chamado por um
    agendador externo (GitHub Actions). Protegido por CRON_SECRET, já que esta
    rota é isenta do login Basic."""
    if not settings.CRON_SECRET:
        raise HTTPException(503, "Cron desativado: defina CRON_SECRET no ambiente.")
    enviado = request.headers.get("X-Cron-Key") or request.query_params.get("chave") or ""
    if not secrets.compare_digest(enviado, settings.CRON_SECRET):
        raise HTTPException(403, "Chave inválida.")
    bg.add_task(_rodar_coleta_bg)
    return {"ok": True, "mensagem": "Coleta iniciada (cron)."}


@app.get("/api/coleta/status")
def coleta_status(user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    """Estado da coleta para o indicador do dashboard.

    "em_andamento"/"travado" é ESTADO GLOBAL (a trava _coleta_lock vale pra
    todo mundo, não só pra este usuário) — usa a MESMA trava que barra o
    botão manual (ver /api/coletar), em vez de inferir pelo último LogColeta
    deste usuário. Achado real: numa coleta de cron (processa vários
    usuários), o registro "terminou" de UM usuário só é criado quando a vez
    dele é concluída — enquanto isso, esse usuário via o botão manual
    recusar com "já existe uma coleta em andamento" (a trava global, correta)
    ao mesmo tempo que o indicador do dashboard mostrava "ocioso" (o LogColeta
    dele ainda não tinha sido tocado nesta rodada) — duas fontes de verdade
    desalinhadas. As estatísticas (novos/vistos/fortes) continuam por
    usuário, já que só ficam prontas quando a rodada chega na conta dele.

    "fase"/"fase_feitos"/"fase_total": em qual etapa a coleta em andamento
    está — "buscando" (baixando/gravando editais do PNCP, uma vez só,
    compartilhado) ou "compatibilidade" (calculando compatibilidade pro
    catálogo de cada usuário — desde a paralelização, "feitos" pode não
    andar na ordem de submissão, os usuários terminam conforme cada um fica
    pronto). Achado real: antes disso, o indicador só dizia "coleta em
    andamento" do início ao fim — parecia uma trava mesmo quando a busca no
    PNCP já tinha terminado fazia tempo e só faltava processar usuários."""
    agora = _utcnow_main()
    em_andamento = _coleta_lock.locked()
    travado = False
    iniciada_ha_seg = None
    if em_andamento and _coleta_iniciada_em is not None:
        iniciada_ha_seg = int((agora - _coleta_iniciada_em).total_seconds())
        if _coleta_travada():
            em_andamento, travado = False, True

    # última coleta concluída DESTE usuário
    ultima_ok = db.execute(
        select(LogColeta).where(LogColeta.usuario_id == user.id,
                                LogColeta.finalizado_em.is_not(None))
        .order_by(LogColeta.id.desc()).limit(1)
    ).scalar_one_or_none()

    if not em_andamento and not travado and not ultima_ok:
        return {"estado": "nunca"}

    estado = "em_andamento" if em_andamento else ("travado" if travado else "ocioso")
    return {
        "estado": estado,
        "iniciada_ha_seg": iniciada_ha_seg,
        "fase": _coleta_fase if em_andamento else None,
        "fase_feitos": _coleta_fase_feitos if em_andamento else None,
        "fase_total": _coleta_fase_total if em_andamento else None,
        "ultima_fim_seg": int((agora - ultima_ok.finalizado_em).total_seconds())
            if ultima_ok and ultima_ok.finalizado_em else None,
        "novos": ultima_ok.editais_novos if ultima_ok else None,
        "vistos": ultima_ok.editais_vistos if ultima_ok else None,
        "fortes": ultima_ok.matches_fortes if ultima_ok else None,
        "erro": ultima_ok.erro if ultima_ok else None,
    }


@app.get("/api/logs")
def logs(user: Usuario = Depends(_auth.get_current_user),
         db: Session = Depends(get_session)):
    regs = db.execute(
        select(LogColeta).where(LogColeta.usuario_id == user.id)
        .order_by(LogColeta.id.desc()).limit(30)
    ).scalars().all()
    return [{
        "id": l.id, "fonte": l.fonte, "origem": l.origem,
        "iniciado_em": _brt(l.iniciado_em),
        "finalizado_em": _brt(l.finalizado_em),
        "editais_novos": l.editais_novos, "editais_vistos": l.editais_vistos,
        "matches_fortes": l.matches_fortes, "erro": l.erro,
    } for l in regs]


@app.get("/api/resumo")
def resumo(user: Usuario = Depends(_auth.get_current_user),
           db: Session = Depends(get_session)):
    hoje = date.today()
    ativo = (Edital.data_abertura.is_(None)) | (Edital.data_abertura >= hoje)
    meu = Match.usuario_id == user.id

    total_prod = db.scalar(
        select(func.count(Produto.id)).where(Produto.usuario_id == user.id)) or 0
    # editais ativos que ESTE usuário tem como match
    total_editais = db.scalar(
        select(func.count(Match.id)).join(Edital, Match.edital_id == Edital.id)
        .where(ativo).where(meu)
    ) or 0
    por_nivel = dict(db.execute(
        select(Match.nivel, func.count(Match.id))
        .join(Edital, Match.edital_id == Edital.id)
        .where(ativo).where(meu)
        .group_by(Match.nivel)
    ).all())
    nao_lidos = db.scalar(
        select(func.count(Match.id))
        .join(Edital, Match.edital_id == Edital.id)
        .where(ativo).where(meu).where(Match.lido == False)  # noqa: E712
    ) or 0
    do_dia, valor_do_dia = db.execute(
        select(func.count(Match.id), func.sum(Edital.valor_estimado))
        .join(Edital, Match.edital_id == Edital.id)
        .where(ativo).where(meu).where(Edital.data_abertura == hoje)
    ).one()
    return {
        "produtos": total_prod, "editais": total_editais,
        "fortes": por_nivel.get("forte", 0), "medios": por_nivel.get("medio", 0),
        "fracos": por_nivel.get("fraco", 0), "nao_lidos": nao_lidos,
        "do_dia": do_dia or 0, "valor_do_dia": valor_do_dia or 0,
    }


@app.get("/api/agenda")
def agenda(offset: int = 0, user: Usuario = Depends(_auth.get_current_user),
          db: Session = Depends(get_session)):
    """Editais com match do usuário cuja janela de propostas abre numa
    semana (domingo a sábado). offset=0 é a semana atual, -1 a anterior, 1 a
    seguinte. Sem o filtro "ativo" que /api/resumo usa: navegar pra uma
    semana passada tem que continuar mostrando o que aconteceu, não some
    só porque o edital já encerrou.

    Usa data_abertura (dataAberturaProposta no PNCP — início do
    recebimento de propostas), por pedido do usuário: o calendário deve
    refletir quando a janela de propostas ABRE, não o prazo final
    (data_encerramento). Isso é deliberadamente diferente de como
    "dias_restantes" era calculado antes desta mudança -- ver o mesmo
    campo em listar_editais/resumo, agora também baseados em data_abertura."""
    hoje = date.today()
    inicio = hoje - timedelta(days=(hoje.weekday() + 1) % 7) + timedelta(weeks=offset)
    fim = inicio + timedelta(days=6)
    linhas = db.execute(
        select(Edital, Match).join(Match, Match.edital_id == Edital.id)
        .where(Match.usuario_id == user.id)
        .where(Edital.data_abertura.between(inicio, fim))
        .order_by(Edital.data_abertura)
    ).all()
    datas_com_sessao = {ed.data_abertura for ed, _m in linhas}
    dias = [{"data": (inicio + timedelta(days=i)).isoformat(),
             "tem_sessao": (inicio + timedelta(days=i)) in datas_com_sessao} for i in range(7)]
    sessoes = [{
        "edital_id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto,
        "modalidade": ed.modalidade, "municipio": ed.municipio, "uf": ed.uf,
        "valor_estimado": ed.valor_estimado, "data_sessao": ed.data_abertura.isoformat(),
    } for ed, _m in linhas]
    return {"inicio": inicio.isoformat(), "fim": fim.isoformat(), "dias": dias, "sessoes": sessoes}


@app.get("/api/editais/recentes")
def editais_recentes(limite: int = 5, user: Usuario = Depends(_auth.get_current_user),
                     db: Session = Depends(get_session)):
    """Últimos editais com que o usuário interagiu -- abriu a página e
    navegou entre as abas (itens/cotação/análise/documentos/proposta), ver
    POST .../interacao -- mais recente primeiro. Alimenta o card
    "Analisados recentemente" do painel Início."""
    limite = max(1, min(limite, 20))
    linhas = db.execute(
        select(Edital, Match).join(Match, Match.edital_id == Edital.id)
        .where(Match.usuario_id == user.id)
        .where(Match.interagido_em.is_not(None))
        .order_by(Match.interagido_em.desc())
        .limit(limite)
    ).all()
    return {"editais": [{
        "edital_id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto,
        "modalidade": ed.modalidade, "municipio": ed.municipio, "uf": ed.uf,
        "valor_estimado": ed.valor_estimado, "nivel": m.nivel, "score": m.score,
        "interagido_em": _brt(m.interagido_em),
    } for ed, m in linhas]}


def _intervalo_mes(ano: int, mes: int) -> tuple[datetime, datetime]:
    inicio = datetime(ano, mes, 1)
    fim = datetime(ano + 1, 1, 1) if mes == 12 else datetime(ano, mes + 1, 1)
    return inicio, fim


def _mes_anterior(ano: int, mes: int) -> tuple[int, int]:
    return (ano - 1, 12) if mes == 1 else (ano, mes - 1)


def _totais_ganhos_mes(db: Session, user: Usuario, ano: int, mes: int) -> dict:
    """Soma valor/custo/margem das Propostas salvas dos editais marcados
    "ganho" nesse mês -- só usada pra comparação com o mês corrente (ver
    GET /api/ganhos), não precisa da lista de editais, só os totais."""
    inicio, fim = _intervalo_mes(ano, mes)
    edital_ids = db.execute(
        select(Match.edital_id).where(
            Match.usuario_id == user.id, Match.status == "ganho",
            Match.status_atualizado_em.is_not(None),
            Match.status_atualizado_em >= inicio, Match.status_atualizado_em < fim,
        )
    ).scalars().all()
    valor_total = custo_total = 0.0
    if edital_ids:
        for p in db.execute(
            select(Proposta).where(Proposta.edital_id.in_(edital_ids), Proposta.usuario_id == user.id)
        ).scalars():
            if not p.itens:
                continue
            valor_total += sum((i.get("preco_unit") or 0) * (i.get("quantidade") or 0) for i in p.itens)
            custo_total += sum((i.get("custo_unit") or 0) * (i.get("quantidade") or 0) for i in p.itens)
    margem_total = valor_total - custo_total
    return {"quantidade": len(edital_ids), "valor_total": round(valor_total, 2),
           "custo_total": round(custo_total, 2), "margem_total": round(margem_total, 2)}


@app.get("/api/ganhos")
def ganhos(ano: int = Query(...), mes: int = Query(..., ge=1, le=12),
          user: Usuario = Depends(_auth.get_current_user),
          db: Session = Depends(get_session)):
    """Editais marcados "ganho" no mês informado -- o mês é quando o status
    virou "ganho" (status_atualizado_em), não nenhuma data do próprio
    edital. Valor/custo/margem vêm da Proposta salva pelo usuário, nunca do
    valor estimado do edital -- edital ganho sem proposta salva conta na
    quantidade mas não entra nos valores (tem_proposta=false). Alimenta o
    card "Editais ganhos" do painel Início, incluindo a comparação com o
    mês anterior."""
    inicio, fim = _intervalo_mes(ano, mes)
    linhas = db.execute(
        select(Match, Edital).join(Edital, Edital.id == Match.edital_id)
        .where(Match.usuario_id == user.id, Match.status == "ganho")
        .where(Match.status_atualizado_em.is_not(None))
        .where(Match.status_atualizado_em >= inicio, Match.status_atualizado_em < fim)
        .order_by(Match.status_atualizado_em.desc())
    ).all()

    edital_ids = [ed.id for _m, ed in linhas]
    propostas = {}
    if edital_ids:
        for p in db.execute(
            select(Proposta).where(Proposta.edital_id.in_(edital_ids), Proposta.usuario_id == user.id)
        ).scalars():
            propostas[p.edital_id] = p

    editais_out = []
    valor_total = custo_total = 0.0
    for _m, ed in linhas:
        prop = propostas.get(ed.id)
        tem_proposta = bool(prop and prop.itens)
        item = {"edital_id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto,
               "modalidade": ed.modalidade, "tem_proposta": tem_proposta}
        if tem_proposta:
            tv = sum((i.get("preco_unit") or 0) * (i.get("quantidade") or 0) for i in prop.itens)
            tc = sum((i.get("custo_unit") or 0) * (i.get("quantidade") or 0) for i in prop.itens)
            margem = tv - tc
            item.update({
                "valor_total": round(tv, 2), "custo_total": round(tc, 2),
                "margem": round(margem, 2),
                "margem_pct": round(margem / tv * 100, 1) if tv else 0.0,
            })
            valor_total += tv
            custo_total += tc
        editais_out.append(item)

    margem_total = valor_total - custo_total
    comparacao = _totais_ganhos_mes(db, user, *_mes_anterior(ano, mes))
    margem_variacao_pct = (round((margem_total - comparacao["margem_total"]) / abs(comparacao["margem_total"]) * 100, 1)
                           if comparacao["margem_total"] else None)

    return {
        "ano": ano, "mes": mes,
        "quantidade": len(linhas),
        "valor_total": round(valor_total, 2), "custo_total": round(custo_total, 2),
        "margem_total": round(margem_total, 2),
        "margem_pct": round(margem_total / valor_total * 100, 1) if valor_total else 0.0,
        "margem_variacao_pct": margem_variacao_pct,
        "editais": editais_out,
    }


_MAX_DIAS_COMPROMISSOS = 62  # ~2 meses -- trava de custo, evita varredura de intervalo absurdo


@app.get("/api/compromissos")
def compromissos(inicio: date, fim: date, user: Usuario = Depends(_auth.get_current_user),
                 db: Session = Depends(get_session)):
    """Calendário de compromissos num intervalo qualquer (semana OU mês —
    quem decide o intervalo é o front, ver abrirCalendarioCompromissos()):
    documentos de habilitação vencendo + editais que o usuário marcou "vou
    participar" abrindo. Junta os dois tipos numa lista só, ordenada por
    data, pra caber num card de calendário sem precisar de duas consultas
    separadas no front.

    Usa data_abertura (dataAberturaProposta no PNCP — início do
    recebimento de propostas) pro edital, por pedido do usuário -- ver
    mesmo motivo em /api/agenda, alguns metros acima."""
    if fim < inicio or (fim - inicio).days > _MAX_DIAS_COMPROMISSOS:
        raise HTTPException(400, "Intervalo inválido.")

    docs = db.execute(
        select(Documento).where(Documento.usuario_id == user.id, Documento.ativo == True,  # noqa: E712
                                Documento.data_validade.between(inicio, fim))
    ).scalars().all()
    editais = db.execute(
        select(Edital).join(Match, Match.edital_id == Edital.id)
        .where(Match.usuario_id == user.id, Match.status == "vou_participar")
        .where(Edital.data_abertura.between(inicio, fim))
    ).scalars().all()

    compromissos_lista = [{
        "tipo": "documento", "data": d.data_validade.isoformat(),
        "documento_id": d.id, "nome": d.nome, "link": d.link,
    } for d in docs] + [{
        "tipo": "edital", "data": ed.data_abertura.isoformat(),
        "edital_id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto,
        "modalidade": ed.modalidade, "municipio": ed.municipio, "uf": ed.uf,
        "valor_estimado": ed.valor_estimado,
    } for ed in editais]
    compromissos_lista.sort(key=lambda c: c["data"])
    dias_com_compromisso = sorted({c["data"] for c in compromissos_lista})
    return {"inicio": inicio.isoformat(), "fim": fim.isoformat(),
            "dias_com_compromisso": dias_com_compromisso, "compromissos": compromissos_lista}


class PropostaIn(BaseModel):
    itens: list[dict] = []
    observacoes: str | None = None


def _proposta_payload(ed: Edital, prop: Proposta | None) -> dict:
    if prop and prop.itens:
        itens = prop.itens
    else:
        # esqueleto a partir dos itens do edital
        itens = [{
            "numero": it.numero,
            "descricao": it.descricao,
            "quantidade": it.quantidade or 0,
            "custo_unit": 0,
            "preco_unit": it.valor_unitario or 0,
        } for it in ed.itens]
    # Achado real: a descrição de um item ficava "congelada" com o que
    # estava em ItemEdital.descricao no momento em que foi adicionado à
    # proposta — se completar_descricao_itens() melhorasse o texto DEPOIS
    # (a API do PNCP vinha cortada, o documento oficial do edital tem a
    # versão completa), uma proposta já salva continuava mostrando/
    # exportando a versão antiga, cortada. Sempre usa a descrição ATUAL de
    # ItemEdital quando o item da proposta tem "numero" (aponta pra um item
    # de verdade do edital) — o resultado é um dict NOVO por item (não
    # mutar prop.itens direto, pra não arriscar persistir sem intenção
    # numa próxima flush da sessão). Item sem "numero" (proposta salva
    # antes dessa referência existir, ou descrição digitada à mão) mantém
    # o texto salvo, sem como atualizar.
    descricoes_atuais = {it.numero: it.descricao for it in ed.itens if it.numero is not None}
    itens = [{**i, "descricao": descricoes_atuais.get(i.get("numero"), i.get("descricao"))} for i in itens]
    total_venda = sum((i.get("preco_unit") or 0) * (i.get("quantidade") or 0) for i in itens)
    total_custo = sum((i.get("custo_unit") or 0) * (i.get("quantidade") or 0) for i in itens)
    margem = total_venda - total_custo
    margem_pct = (margem / total_venda * 100) if total_venda else 0
    # itens do edital (TODOS, não só os já incluídos na proposta) — usado
    # pelo front pra montar o modal de "adicionar item", restrito ao que o
    # edital de fato pede, em vez de deixar digitar qualquer coisa.
    itens_edital = [{
        "numero": it.numero, "descricao": it.descricao,
        "quantidade": it.quantidade or 0, "valor_unitario": it.valor_unitario or 0,
    } for it in ed.itens]
    return {
        "edital_id": ed.id, "orgao": ed.orgao, "objeto": ed.objeto,
        "itens": itens, "itens_edital": itens_edital,
        "observacoes": prop.observacoes if prop else "",
        "total_venda": round(total_venda, 2), "total_custo": round(total_custo, 2),
        "margem": round(margem, 2), "margem_pct": round(margem_pct, 1),
        "existe": prop is not None,
    }


@app.get("/api/editais/{edital_id}/proposta")
def obter_proposta(edital_id: int, user: Usuario = Depends(_auth.get_current_user),
                   db: Session = Depends(get_session)):
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    prop = db.execute(select(Proposta).where(Proposta.edital_id == edital_id)
                      .where(Proposta.usuario_id == user.id)).scalars().first()
    return _proposta_payload(ed, prop)


@app.post("/api/editais/{edital_id}/proposta")
def salvar_proposta(edital_id: int, dados: PropostaIn,
                    user: Usuario = Depends(_auth.get_current_user),
                    db: Session = Depends(get_session)):
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    prop = db.execute(select(Proposta).where(Proposta.edital_id == edital_id)
                      .where(Proposta.usuario_id == user.id)).scalars().first()
    if prop is None:
        prop = Proposta(edital_id=edital_id, usuario_id=user.id)
        db.add(prop)
    prop.itens = dados.itens
    prop.observacoes = dados.observacoes
    db.commit()
    db.refresh(prop)
    return _proposta_payload(ed, prop)


def _dados_remetente(user: Usuario) -> dict:
    """Dados do PRÓPRIO usuário (proponente) pra timbrar a proposta —
    decifra endereço/dados complementares, igual a /api/perfil."""
    import json as _j
    end = _auth.decifrar(user.endereco_cifrado)
    try:
        endereco = _j.loads(end) if end else {}
    except ValueError:
        endereco = {}
    emp = _auth.decifrar(user.dados_empresa_cifrado)
    try:
        empresa = _j.loads(emp) if emp else {}
    except ValueError:
        empresa = {}
    return {
        "nome": user.nome, "documento": _auth.decifrar(user.doc_cifrado),
        "endereco": endereco, "empresa": empresa, "logo_base64": user.logo_base64,
    }


@app.get("/api/editais/{edital_id}/proposta.pdf")
def exportar_proposta_pdf(edital_id: int, user: Usuario = Depends(_auth.get_current_user),
                          db: Session = Depends(get_session)):
    ed = db.get(Edital, edital_id)
    if not ed:
        raise HTTPException(404, "Edital não encontrado")
    prop = db.execute(select(Proposta).where(Proposta.edital_id == edital_id)
                      .where(Proposta.usuario_id == user.id)).scalars().first()
    p = _proposta_payload(ed, prop)
    edital_info = {
        "orgao": ed.orgao, "objeto": ed.objeto, "modalidade": ed.modalidade,
        "municipio": ed.municipio, "uf": ed.uf, "id_externo": ed.id_externo,
        "data_encerramento": ed.data_encerramento.isoformat() if ed.data_encerramento else None,
        "link": ed.link,
    }
    from .proposta_pdf import gerar_pdf_proposta
    pdf_bytes = gerar_pdf_proposta(_dados_remetente(user), edital_info, p)
    nome = f"proposta_edital_{edital_id}.pdf"
    return StreamingResponse(iter([pdf_bytes]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={nome}"})


@app.get("/api/export.csv")
def export_csv(nivel: str | None = None,
               user: Usuario = Depends(_auth.get_current_user),
               db: Session = Depends(get_session)):
    q = select(Match, Edital).join(Edital, Match.edital_id == Edital.id) \
        .where(Match.usuario_id == user.id)
    if nivel:
        q = q.where(Match.nivel == nivel)
    q = q.order_by(Match.score.desc())

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["nivel", "score", "itens_compativeis", "orgao", "uf", "municipio",
                "modalidade", "valor_estimado", "data_encerramento", "objeto", "link"])
    for m, ed in db.execute(q).all():
        w.writerow([m.nivel, m.score, m.itens_compativeis, ed.orgao, ed.uf,
                    ed.municipio, ed.modalidade, ed.valor_estimado,
                    ed.data_encerramento, (ed.objeto or "")[:300], ed.link])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=editais.csv"},
    )


# --------------------------- Catálogo CATMAT/CATSER ------------------- #
@app.get("/api/catmat")
def buscar_catmat(
    descricao: str = Query(..., min_length=2),
    tipo: str = Query("material", pattern="^(material|servico)$"),
    debug: bool = Query(False),
    user: Usuario = Depends(_auth.get_current_user),
):
    """Busca códigos CATMAT (material) ou CATSER (serviço) na API oficial
    de dados abertos do Compras.gov.br, ranqueados por relevância.
    Use ?debug=true para diagnosticar o que a API externa devolveu."""
    r = catmat.buscar(descricao, tipo=tipo, debug=debug)
    saida = {"status": r["status"], "total": len(r["itens"]), "resultados": r["itens"]}
    if "debug" in r:
        saida["debug"] = r["debug"]
    return saida


# --------------------------- Documentos (habilitação) ----------------- #
@app.get("/api/documentos")
def listar_documentos(user: Usuario = Depends(_auth.get_current_user),
                      db: Session = Depends(get_session)):
    docs = db.execute(select(Documento).where(Documento.usuario_id == user.id)
                      .order_by(Documento.data_validade.asc())).scalars().all()
    hoje = date.today()
    return [{
        "id": d.id, "nome": d.nome, "orgao_emissor": d.orgao_emissor,
        "data_validade": d.data_validade.isoformat(),
        "dias_para_vencer": (d.data_validade - hoje).days,
        "observacao": d.observacao, "link": d.link, "ativo": d.ativo,
        "tem_arquivo": bool(d.arquivo_cifrado),
    } for d in docs]


_TIPOS_UPLOAD_DOCUMENTO_PERMITIDOS = {"application/pdf", "image/jpeg", "image/png", "image/webp"}
_TAMANHO_MAX_UPLOAD_DOCUMENTO = 15 * 1024 * 1024  # 15 MB


async def _ler_upload_documento(arquivo: UploadFile | None) -> bytes:
    """Valida e lê os bytes de um upload de documento de habilitação. O
    arquivo é obrigatório -- é o "cofre" propriamente dito (o que fica
    guardado e pode ser baixado depois), não só uma fonte de texto pra IA."""
    if arquivo is None or not arquivo.filename:
        raise HTTPException(400, "Envie o arquivo do documento (PDF ou imagem).")
    if arquivo.content_type not in _TIPOS_UPLOAD_DOCUMENTO_PERMITIDOS:
        raise HTTPException(400, "Envie um PDF ou imagem (jpg/png/webp).")
    conteudo = await arquivo.read()
    if len(conteudo) > _TAMANHO_MAX_UPLOAD_DOCUMENTO:
        raise HTTPException(400, "Arquivo muito grande (máximo 15 MB).")
    return conteudo


def _nome_arquivo_seguro(nome: str) -> str:
    """Mesmo saneamento já usado em /api/documentos/baixar (linha ~2138):
    tira caracteres que quebrariam o header Content-Disposition."""
    seguro = re.sub(r'[\\/:*?"<>|\r\n]', "_", nome or "documento").strip()[:150] or "documento"
    return seguro.encode("latin-1", errors="ignore").decode("latin-1") or "documento"


@app.post("/api/documentos")
async def criar_documento(nome: str = Form(...), orgao_emissor: str | None = Form(None),
                          data_validade: date | None = Form(None), link: str | None = Form(None),
                          observacao: str | None = Form(None),
                          arquivo: UploadFile = File(...),
                          user: Usuario = Depends(_auth.get_current_user),
                          db: Session = Depends(get_session)):
    conteudo = await _ler_upload_documento(arquivo)
    from . import analise_edital as ia
    texto = ia.extrair_texto_upload(arquivo.filename, conteudo, arquivo.content_type) or None

    # v1 do cofre: a IA só extrai a data de validade quando o usuário não
    # digitou uma -- nenhum julgamento de "atende exigência X" ou apto/
    # inapto acontece aqui (isso é outra funcionalidade, verificar_documentos
    # _usuario, que fica intocada). Sem data segura, não inventa: 422 pede
    # pro usuário confirmar manualmente.
    if data_validade is None:
        chave = _auth.decifrar(user.gemini_key_cifrada)
        data_validade = ia.extrair_validade_documento(texto or "", api_key=chave)
        if data_validade is None:
            raise HTTPException(
                422, "Não foi possível identificar a validade automaticamente. Informe a data manualmente.")

    d = Documento(nome=nome, orgao_emissor=orgao_emissor or None, data_validade=data_validade,
                 link=link or None, observacao=observacao or None, texto_extraido=texto,
                 arquivo_cifrado=_auth.cifrar(base64.b64encode(conteudo).decode("ascii")),
                 arquivo_nome=arquivo.filename, arquivo_tipo=arquivo.content_type,
                 usuario_id=user.id)
    db.add(d)
    user.versao_documentos += 1
    db.commit()
    return {"id": d.id, "data_validade": d.data_validade.isoformat()}


def _documento_do_usuario(db, doc_id, user) -> Documento:
    d = db.get(Documento, doc_id)
    if not d or d.usuario_id != user.id:
        raise HTTPException(404, "Documento não encontrado")
    return d


@app.get("/api/documentos/{doc_id}/arquivo")
def baixar_arquivo_documento(doc_id: int, user: Usuario = Depends(_auth.get_current_user),
                             db: Session = Depends(get_session)):
    d = _documento_do_usuario(db, doc_id, user)   # 404 se não for do usuário logado
    if not d.arquivo_cifrado:
        raise HTTPException(404, "Este documento não tem arquivo salvo.")
    conteudo = base64.b64decode(_auth.decifrar(d.arquivo_cifrado))
    return Response(content=conteudo, media_type=d.arquivo_tipo or "application/octet-stream",
                    headers={"Content-Disposition":
                             f'attachment; filename="{_nome_arquivo_seguro(d.arquivo_nome)}"'})


@app.put("/api/documentos/{doc_id}")
async def atualizar_documento(doc_id: int, nome: str = Form(...), orgao_emissor: str | None = Form(None),
                              data_validade: date = Form(...), link: str | None = Form(None),
                              observacao: str | None = Form(None),
                              arquivo: UploadFile | None = File(None),
                              user: Usuario = Depends(_auth.get_current_user),
                              db: Session = Depends(get_session)):
    d = _documento_do_usuario(db, doc_id, user)
    d.nome, d.orgao_emissor = nome, orgao_emissor or None
    d.data_validade, d.link, d.observacao = data_validade, link or None, observacao or None
    # arquivo é OPCIONAL na edição: só substitui o arquivo/texto salvos se o
    # usuário mandar um novo; sem arquivo, mantém o que já estava guardado
    # (editar validade/observação não pode apagar o upload anterior).
    if arquivo is not None and arquivo.filename:
        conteudo = await _ler_upload_documento(arquivo)
        from . import analise_edital as ia
        d.texto_extraido = ia.extrair_texto_upload(arquivo.filename, conteudo, arquivo.content_type) or None
        d.arquivo_cifrado = _auth.cifrar(base64.b64encode(conteudo).decode("ascii"))
        d.arquivo_nome, d.arquivo_tipo = arquivo.filename, arquivo.content_type
    d.avisado_para = None  # validade mudou -> permite avisar de novo
    user.versao_documentos += 1
    db.commit()
    return {"ok": True}


@app.delete("/api/documentos/{doc_id}")
def remover_documento(doc_id: int, user: Usuario = Depends(_auth.get_current_user),
                      db: Session = Depends(get_session)):
    d = _documento_do_usuario(db, doc_id, user)
    db.delete(d)
    user.versao_documentos += 1
    db.commit()
    return {"ok": True}


@app.post("/api/lembretes/verificar")
def verificar_lembretes(bg: BackgroundTasks):
    """Dispara a verificação de prazos e documentos manualmente."""
    def _run():
        db = SessionLocal()
        try:
            from .lembretes import verificar_todos
            verificar_todos(db)
        finally:
            db.close()
    bg.add_task(_run)
    return {"ok": True, "mensagem": "Verificação de lembretes iniciada."}


# --------------------------- Configurações ---------------------------- #
class ConfigIn(BaseModel):
    PNCP_UFS: str | None = None
    PNCP_MODALIDADES: str | None = None
    PNCP_HORIZONTE_DIAS: str | None = None
    IA_ATIVA: str | None = None


@app.get("/api/config")
def obter_config(user: Usuario = Depends(_auth.get_current_user),
                 db: Session = Depends(get_session)):
    from . import configuracoes
    from .matching.embeddings import ia_disponivel, ia_bloqueada, segundos_para_liberar
    dados = configuracoes.todas(db)
    chave_user = _auth.decifrar(user.gemini_key_cifrada)
    dados["IA_DISPONIVEL"] = "1" if ia_disponivel(chave_user) else "0"  # chave (do user ou global)?
    dados["IA_CHAVE_PROPRIA"] = "1" if chave_user else "0"             # usa chave própria?
    dados["IA_BLOQUEADA"] = "1" if ia_bloqueada(chave_user) else "0"  # cota diária ESTOUROU PRA ESSE USUÁRIO?
    dados["IA_LIBERA_EM_MIN"] = str(round(segundos_para_liberar(chave_user) / 60))
    return dados


@app.post("/api/config")
def salvar_config(dados: ConfigIn, user: Usuario = Depends(_auth.get_current_user),
                  db: Session = Depends(get_session)):
    from . import configuracoes
    for chave, valor in dados.model_dump().items():
        if valor is not None:
            configuracoes.definir(db, chave, valor.strip())
    return {"ok": True, "config": configuracoes.todas(db)}


# --------------------------- Inteligência de preço -------------------- #
# Detecta quando a descrição do item deixa claro que a unidade cotada é uma
# EMBALAGEM com várias peças dentro (ex.: "caixa com 250 unidades", "(100
# UND)") — sem isso, a Inteligência de Preço comparava "preço de 1 peça
# avulsa" com "preço de uma caixa de 250" como se fossem a mesma grandeza,
# distorcendo mediana/mínimo/máximo em até 100x+. Achado real em produção:
# "Envelope Kraft" variava de R$0,84 a R$164,80 só por causa dessa mistura
# de escala entre editais que cotam por unidade e editais que cotam por caixa.
_RE_MULTIPLICADOR_EMBALAGEM = re.compile(
    r"(?:caixa|cx|pacote|pct|kit|fardo|embalagem)\s*(?:com|c/)?\s*(?P<n1>\d+)\s*"
    r"(?:unidades?|und?|un\.?|pe[cç]as?|folhas?)?\b"
    r"|(?:com|c/)\s*(?P<n2>\d+)\s*(?:unidades?|und?|un\.?|pe[cç]as?|folhas?)\b"
    r"|\(\s*(?P<n3>\d+)\s*(?:unidades?|und?|un\.?)\s*\)",
    re.IGNORECASE,
)


def _valor_unitario_normalizado(item: ItemEdital) -> float | None:
    """Valor unitário do item, dividido pelo multiplicador de embalagem
    quando a descrição deixa claro que a unidade cotada é uma caixa/pacote
    com várias peças — pra comparar "preço por peça" com "preço por peça"
    na Inteligência de Preço, não misturar escalas diferentes. Sem isso, um
    item cotado "caixa com 250" e outro cotado por unidade avulsa entravam
    na mesma amostra como se fossem o mesmo tipo de valor."""
    if item.valor_unitario is None or item.valor_unitario <= 0:
        return None
    m = _RE_MULTIPLICADOR_EMBALAGEM.search(item.descricao or "")
    if m:
        n = int(m.group("n1") or m.group("n2") or m.group("n3"))
        if n > 1:
            return item.valor_unitario / n
    return item.valor_unitario


def _banda_outlier_preco(valores: list[float], multiplicador: float = 15,
                         amostra_minima: int = 5) -> tuple[float, float] | None:
    """Faixa [mediana_bruta/multiplicador, mediana_bruta*multiplicador] usada
    pra reconhecer valor unitário absurdo (normalmente erro de digitação do
    órgão no PNCP — ex.: valor total do lote lançado no campo de valor
    unitário). None = amostra pequena demais pra mediana bruta ser confiável
    (com poucas ocorrências, uma variação real de preço podia ser confundida
    com erro)."""
    if len(valores) < amostra_minima:
        return None
    ordenados = sorted(valores)
    n = len(ordenados)
    mediana_bruta = ordenados[n // 2] if n % 2 else (ordenados[n // 2 - 1] + ordenados[n // 2]) / 2
    if mediana_bruta <= 0:
        return None
    return mediana_bruta / multiplicador, mediana_bruta * multiplicador


def _filtrar_outliers_preco(valores: list[float]) -> list[float]:
    banda = _banda_outlier_preco(valores)
    if banda is None:
        return valores
    limite_inf, limite_sup = banda
    filtrados = [v for v in valores if limite_inf <= v <= limite_sup]
    return filtrados or valores   # nunca esvazia a amostra por engano


@app.get("/api/inteligencia-preco")
def inteligencia_preco(user: Usuario = Depends(_auth.get_current_user),
                       db: Session = Depends(get_session)):
    """Para cada produto, estatísticas do valor unitário estimado pelo órgão
    nos itens de edital em que ele apareceu como compatível. Dá uma referência
    de mercado com base no histórico que o próprio sistema acumulou.

    Obs.: usa o valor unitário ESTIMADO do item (não o preço homologado do
    vencedor — isso exigiria puxar os resultados/atas do PNCP, um passo
    futuro). Importante: é o valor do ITEM, não o valor total do edital —
    um edital pode ter dezenas de itens somando milhões, mas o item que
    casou com este produto pode valer poucos reais."""
    produtos = db.execute(select(Produto).where(Produto.usuario_id == user.id)).scalars().all()
    matches = db.execute(
        select(Match).where(Match.usuario_id == user.id, Match.detalhe.is_not(None))
    ).scalars().all()

    # (edital_id, número do item) -> produto_id, a partir do detalhe de cada match
    numeros_por_edital: dict[int, set] = {}
    referencias: list[tuple[int, int, int]] = []  # (edital_id, numero, produto_id)
    for m in matches:
        for d in (m.detalhe or {}).get("itens", []):
            numero, produto_id = d.get("item"), d.get("produto_id")
            if numero is None or produto_id is None:
                continue
            # só entra na estatística o que é confiável (código exato/score
            # alto) ou confirmado manualmente — sugestão não confirmada não
            # vira "referência de mercado".
            if not (d.get("confianca") == "alta" or d.get("confirmado_manualmente")):
                continue
            referencias.append((m.edital_id, numero, produto_id))
            numeros_por_edital.setdefault(m.edital_id, set()).add(numero)

    # busca de uma vez o valor unitário de todos os itens referenciados —
    # valor <= 0 conta como ausente (órgão não preencheu; PNCP manda 0 em vez
    # de vazio), senão o "mínimo" da referência fica sempre zerado à toa.
    valor_do_item: dict[tuple[int, int], float] = {}
    if numeros_por_edital:
        for it in db.execute(
            select(ItemEdital).where(ItemEdital.edital_id.in_(numeros_por_edital.keys()))
        ).scalars():
            if it.numero not in numeros_por_edital.get(it.edital_id, ()):
                continue
            v = _valor_unitario_normalizado(it)
            if v is not None:
                valor_do_item[(it.edital_id, it.numero)] = v

    valores_por_produto: dict[int, list[float]] = {}
    for edital_id, numero, produto_id in referencias:
        v = valor_do_item.get((edital_id, numero))
        if v is not None:
            valores_por_produto.setdefault(produto_id, []).append(v)

    saida = []
    for p in produtos:
        valores = valores_por_produto.get(p.id)
        if not valores:
            continue
        valores = _filtrar_outliers_preco(valores)
        valores.sort()
        n = len(valores)
        mediana = valores[n // 2] if n % 2 else (valores[n // 2 - 1] + valores[n // 2]) / 2
        saida.append({
            "produto_id": p.id, "descricao": p.descricao,
            "ocorrencias": n,
            "minimo": round(min(valores), 2),
            "mediana": round(mediana, 2),
            "media": round(sum(valores) / n, 2),
            "maximo": round(max(valores), 2),
            "preco_venda": p.preco_venda,
        })
    saida.sort(key=lambda x: x["ocorrencias"], reverse=True)
    return saida


@app.get("/api/inteligencia-preco/{produto_id}/editais")
def inteligencia_preco_editais(produto_id: int, user: Usuario = Depends(_auth.get_current_user),
                               db: Session = Depends(get_session)):
    """Lista os editais que embasam as estatísticas de um produto na
    Inteligência de preço, pra conferir de onde vêm os números (inclui os
    que foram descartados do cálculo, com o motivo)."""
    _produto_do_usuario(db, produto_id, user)   # 404 se o produto não é do usuário
    matches = db.execute(
        select(Match).where(Match.usuario_id == user.id, Match.detalhe.is_not(None))
    ).scalars().all()

    referencias = []   # (edital_id, numero do item)
    for m in matches:
        for d in (m.detalhe or {}).get("itens", []):
            if (d.get("produto_id") == produto_id and d.get("item") is not None
                    and (d.get("confianca") == "alta" or d.get("confirmado_manualmente"))):
                referencias.append((m.edital_id, d["item"]))
    if not referencias:
        return []

    edital_ids = {eid for eid, _ in referencias}
    editais = {e.id: e for e in db.execute(
        select(Edital).where(Edital.id.in_(edital_ids))).scalars()}
    itens_map = {(it.edital_id, it.numero): it for it in db.execute(
        select(ItemEdital).where(ItemEdital.edital_id.in_(edital_ids))).scalars()}

    valores_validos = [v for it in itens_map.values()
                       if (v := _valor_unitario_normalizado(it)) is not None]
    banda = _banda_outlier_preco(valores_validos)

    linhas = []
    for edital_id, numero in referencias:
        it = itens_map.get((edital_id, numero))
        ed = editais.get(edital_id)
        if not it or not ed:
            continue
        valor = _valor_unitario_normalizado(it)
        tem_valor = valor is not None
        motivo_exclusao = None
        if not tem_valor:
            motivo_exclusao = "sem_valor"
        elif banda is not None and not (banda[0] <= valor <= banda[1]):
            motivo_exclusao = "fora_do_padrao"
        linhas.append({
            "edital_id": edital_id, "orgao": ed.orgao,
            "municipio": ed.municipio, "uf": ed.uf,
            "descricao_item": it.descricao, "valor_unitario": valor,
            "valor_original": it.valor_unitario if valor != it.valor_unitario else None,
            "link": ed.link, "usado_no_calculo": motivo_exclusao is None,
            "motivo_exclusao": motivo_exclusao,
        })
    linhas.sort(key=lambda x: (not x["usado_no_calculo"], -(x["valor_unitario"] or 0)))
    return linhas


# --------------------------- Dashboard estático ----------------------- #
@app.get("/")
def index():
    return FileResponse(
        os.path.join(STATIC_DIR, "index.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/edital/{edital_id}")
def pagina_edital(edital_id: int):
    """Página própria de um edital (itens, documentos, análise por IA,
    proposta). Mesmo dashboard servido em '/' — o JS decide o que mostrar
    lendo location.pathname (mesmo padrão de /login e afins abaixo)."""
    return FileResponse(
        os.path.join(STATIC_DIR, "index.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/login")
@app.get("/cadastro")
@app.get("/verificar")
@app.get("/redefinir-senha")
def pagina_login():
    """Página única de login/cadastro/verificação/redefinição (decide pela URL no JS)."""
    return FileResponse(
        os.path.join(STATIC_DIR, "login.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
